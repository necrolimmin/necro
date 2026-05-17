from calendar import monthrange
from datetime import date, datetime, timedelta

from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import StationDailyTable1
from accounts.models import (
    KvartalniyGroupExtraPlan,
    KvartalniyMonthly,
    KvartalniyMonthlyPlan,
    StationProfile,
)
from reports.kvartalniy import DISPLAY_GROUPS, _safe_int
import re
from playwright.sync_api import sync_playwright
from django.urls import reverse


def _safe_date(date_str, fallback):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else fallback
    except (TypeError, ValueError):
        return fallback


def _month_start(dt: date) -> date:
    return dt.replace(day=1)


def _month_end(dt: date) -> date:
    return dt.replace(day=monthrange(dt.year, dt.month)[1])


def _same_day_last_year(dt: date) -> date:
    try:
        return dt.replace(year=dt.year - 1)
    except ValueError:
        return dt.replace(year=dt.year - 1, day=28)


def _iter_month_starts(from_date: date, to_date: date):
    cur = from_date.replace(day=1)
    last = to_date.replace(day=1)

    while cur <= last:
        yield cur
        if cur.month == 12:
            cur = cur.replace(year=cur.year + 1, month=1, day=1)
        else:
            cur = cur.replace(month=cur.month + 1, day=1)


def _build_date_range(from_date: date, to_date: date) -> list[date]:
    days = []
    cur = from_date
    while cur <= to_date:
        days.append(cur)
        cur = cur + timedelta(days=1)
    return days


def _get_station_profile_from_user(user):
    return StationProfile.objects.get(user=user)


def _normalize_station_name(value: str) -> str:
    if not value:
        return ""
    return (
        str(value)
        .strip()
        .replace("’", "'")
        .replace("‘", "'")
        .replace("`", "'")
        .replace("ʻ", "'")
        .replace("ʼ", "'")
        .replace("  ", " ")
        .lower()
    )


def _payload_to_metrics(payload: dict) -> dict:
    payload = payload or {}
    return {
        "pogr": _safe_int(payload.get("pogr_itogo", 0)),
        "vygr": _safe_int(payload.get("vygr_itogo", 0)),
        "pogr_kont": _safe_int(payload.get("pogr_itogo_kon", 0)),
        "vygr_kont": _safe_int(payload.get("vygr_itogo_kon", 0)),
        "income": _safe_int(payload.get("income_daily", 0)),
    }


def _aggregate_table1_by_station(date_list: list[date]) -> dict:
    if not date_list:
        return {}

    qs = (
        StationDailyTable1.objects
        .filter(date__in=date_list)
        .select_related("station_user")
        .exclude(shift="total")
    )

    station_map = {}

    for obj in qs:
        try:
            station = _get_station_profile_from_user(obj.station_user)
        except StationProfile.DoesNotExist:
            continue

        metrics = _payload_to_metrics(obj.data or {})

        if station.id not in station_map:
            station_map[station.id] = {
                "station": station,
                "pogr": 0,
                "vygr": 0,
                "pogr_kont": 0,
                "vygr_kont": 0,
                "income": 0,
            }

        station_map[station.id]["pogr"] += metrics["pogr"]
        station_map[station.id]["vygr"] += metrics["vygr"]
        station_map[station.id]["pogr_kont"] += metrics["pogr_kont"]
        station_map[station.id]["vygr_kont"] += metrics["vygr_kont"]
        station_map[station.id]["income"] += metrics["income"]

    return station_map


def _sum_scaled_plans_between(from_date: date, to_date: date):
    result = {}
    month_info = []
    all_full_months = True

    for month_dt in _iter_month_starts(from_date, to_date):
        m_start = _month_start(month_dt)
        m_end = _month_end(month_dt)

        piece_start = max(from_date, m_start)
        piece_end = min(to_date, m_end)

        selected_days = (piece_end - piece_start).days + 1
        days_in_month = monthrange(month_dt.year, month_dt.month)[1]
        is_full_month = piece_start == m_start and piece_end == m_end

        if not is_full_month:
            all_full_months = False

        monthly_obj, _ = KvartalniyMonthly.objects.get_or_create(date=m_start)

        plans = (
            KvartalniyMonthlyPlan.objects
            .filter(monthly=monthly_obj)
            .select_related("station")
        )

        month_info.append({
            "month": m_start,
            "piece_start": piece_start,
            "piece_end": piece_end,
            "selected_days": selected_days,
            "days_in_month": days_in_month,
            "is_full_month": is_full_month,
        })

        for p in plans:
            station = p.station

            if station.id not in result:
                result[station.id] = {
                    "station": station,
                    "pogr_plan": 0,
                    "vygr_plan": 0,
                    "pogr_kont_plan": 0,
                    "vygr_kont_plan": 0,
                    "income_plan": 0,
                }

            result[station.id]["pogr_plan"] += round((p.pogr_plan or 0) / days_in_month * selected_days)
            result[station.id]["vygr_plan"] += round((p.vygr_plan or 0) / days_in_month * selected_days)
            result[station.id]["pogr_kont_plan"] += round((p.pogr_kont_plan or 0) / days_in_month * selected_days)
            result[station.id]["vygr_kont_plan"] += round((p.vygr_kont_plan or 0) / days_in_month * selected_days)
            result[station.id]["income_plan"] += round((p.income_plan or 0) / days_in_month * selected_days)

    return result, month_info, all_full_months


def _row_to_range_dict(station, fact_this=None, fact_last=None, plan_data=None):
    fact_this = fact_this or {}
    fact_last = fact_last or {}
    plan_data = plan_data or {}

    pogr_this = fact_this.get("pogr", 0)
    vygr_this = fact_this.get("vygr", 0)
    pogr_kont_this = fact_this.get("pogr_kont", 0)
    vygr_kont_this = fact_this.get("vygr_kont", 0)
    income_this = fact_this.get("income", 0)

    pogr_last = fact_last.get("pogr", 0)
    vygr_last = fact_last.get("vygr", 0)
    pogr_kont_last = fact_last.get("pogr_kont", 0)
    vygr_kont_last = fact_last.get("vygr_kont", 0)
    income_last = fact_last.get("income", 0)

    return {
        "station_id": station.id if station else None,
        "station_name": station.station_name if station else "",
        "pogr_plan": plan_data.get("pogr_plan", 0),
        "vygr_plan": plan_data.get("vygr_plan", 0),
        "pogr_kont_plan": plan_data.get("pogr_kont_plan", 0),
        "vygr_kont_plan": plan_data.get("vygr_kont_plan", 0),
        "income_plan": plan_data.get("income_plan", 0),

        "pogr_this_year": pogr_this,
        "pogr_last_year": pogr_last,
        "pogr_diff": pogr_this - pogr_last,

        "vygr_this_year": vygr_this,
        "vygr_last_year": vygr_last,
        "vygr_diff": vygr_this - vygr_last,

        "pogr_kont_this_year": pogr_kont_this,
        "pogr_kont_last_year": pogr_kont_last,
        "pogr_kont_diff": pogr_kont_this - pogr_kont_last,

        "vygr_kont_this_year": vygr_kont_this,
        "vygr_kont_last_year": vygr_kont_last,
        "vygr_kont_diff": vygr_kont_this - vygr_kont_last,

        "income_this_year": income_this,
        "income_last_year": income_last,
        "income_diff": income_this - income_last,
    }


def _make_empty_range_row(station_name):
    dummy_station = type("DummyStation", (), {"id": None, "station_name": station_name})()
    return _row_to_range_dict(dummy_station, {}, {}, {})


def _make_zero_totals(label):
    return {
        "station_name": label,
        "pogr_plan": 0,
        "vygr_plan": 0,
        "pogr_kont_plan": 0,
        "vygr_kont_plan": 0,
        "income_plan": 0,

        "pogr_this_year": 0,
        "pogr_last_year": 0,
        "pogr_diff": 0,

        "vygr_this_year": 0,
        "vygr_last_year": 0,
        "vygr_diff": 0,

        "pogr_kont_this_year": 0,
        "pogr_kont_last_year": 0,
        "pogr_kont_diff": 0,

        "vygr_kont_this_year": 0,
        "vygr_kont_last_year": 0,
        "vygr_kont_diff": 0,

        "income_this_year": 0,
        "income_last_year": 0,
        "income_diff": 0,
    }


def _add_to_totals(target, row):
    for key in [
        "pogr_plan", "vygr_plan", "pogr_kont_plan", "vygr_kont_plan", "income_plan",
        "pogr_this_year", "pogr_last_year",
        "vygr_this_year", "vygr_last_year",
        "pogr_kont_this_year", "pogr_kont_last_year",
        "vygr_kont_this_year", "vygr_kont_last_year",
        "income_this_year", "income_last_year",
    ]:
        target[key] += row.get(key, 0) or 0

    target["pogr_diff"] = target["pogr_this_year"] - target["pogr_last_year"]
    target["vygr_diff"] = target["vygr_this_year"] - target["vygr_last_year"]
    target["pogr_kont_diff"] = target["pogr_kont_this_year"] - target["pogr_kont_last_year"]
    target["vygr_kont_diff"] = target["vygr_kont_this_year"] - target["vygr_kont_last_year"]
    target["income_diff"] = target["income_this_year"] - target["income_last_year"]


def _build_kvartalniy_range_context(from_date, to_date):
    if from_date > to_date:
        from_date, to_date = to_date, from_date

    prev_from_date = _same_day_last_year(from_date)
    prev_to_date = _same_day_last_year(to_date)

    selected_dates = _build_date_range(from_date, to_date)
    prev_selected_dates = [_same_day_last_year(d) for d in selected_dates]

    current_data = _aggregate_table1_by_station(selected_dates)
    last_year_data = _aggregate_table1_by_station(prev_selected_dates)

    scaled_plans, month_info, all_full_months = _sum_scaled_plans_between(from_date, to_date)

    single_full_month = all_full_months and len(month_info) == 1
    target_month = month_info[0]["month"] if single_full_month else None

    def _sum_veshoz_between():
        result = {}

        for info in month_info:
            month_date = info["month"]
            selected_days = info["selected_days"]
            days_in_month = info["days_in_month"]

            monthly_obj, _ = KvartalniyMonthly.objects.get_or_create(date=month_date)

            extra_qs = KvartalniyGroupExtraPlan.objects.filter(
                monthly=monthly_obj,
                row_name="Boshqa Stansiya",
            )

            for extra_obj in extra_qs:
                group_key = extra_obj.group_key

                if group_key not in result:
                    result[group_key] = {
                        "station_id": None,
                        "station_name": "Boshqa Stansiya",
                        "is_other": False,
                        "is_veshoz": True,
                        "group_key": group_key,
                        "is_editable": bool(single_full_month),

                        "pogr_plan": 0,
                        "vygr_plan": 0,
                        "pogr_kont_plan": 0,
                        "vygr_kont_plan": 0,

                        "pogr_plan_raw": 0,
                        "vygr_plan_raw": 0,
                        "pogr_kont_plan_raw": 0,
                        "vygr_kont_plan_raw": 0,

                        "pogr_this_year": 0,
                        "pogr_last_year": 0,
                        "pogr_diff": 0,

                        "vygr_this_year": 0,
                        "vygr_last_year": 0,
                        "vygr_diff": 0,

                        "pogr_kont_this_year": 0,
                        "pogr_kont_last_year": 0,
                        "pogr_kont_diff": 0,

                        "vygr_kont_this_year": 0,
                        "vygr_kont_last_year": 0,
                        "vygr_kont_diff": 0,

                        "pogr_this_year_raw": 0,
                        "pogr_last_year_raw": 0,
                        "vygr_this_year_raw": 0,
                        "vygr_last_year_raw": 0,
                        "pogr_kont_this_year_raw": 0,
                        "pogr_kont_last_year_raw": 0,
                        "vygr_kont_this_year_raw": 0,
                        "vygr_kont_last_year_raw": 0,

                        "income_plan": 0,
                        "income_this_year": 0,
                        "income_last_year": 0,
                        "income_diff": 0,

                        "income_plan_raw": 0,
                        "income_this_year_raw": 0,
                        "income_last_year_raw": 0,
                    }

                if single_full_month and month_date == target_month:
                    result[group_key]["pogr_plan_raw"] = getattr(extra_obj, "pogr_plan", 0) or 0
                    result[group_key]["vygr_plan_raw"] = getattr(extra_obj, "vygr_plan", 0) or 0
                    result[group_key]["pogr_kont_plan_raw"] = getattr(extra_obj, "pogr_kont_plan", 0) or 0
                    result[group_key]["vygr_kont_plan_raw"] = getattr(extra_obj, "vygr_kont_plan", 0) or 0
                    result[group_key]["income_plan_raw"] = getattr(extra_obj, "income_plan", 0) // 1000 or 0

                    result[group_key]["pogr_this_year_raw"] = getattr(extra_obj, "pogr_this_year", 0) or 0
                    result[group_key]["pogr_last_year_raw"] = getattr(extra_obj, "pogr_last_year", 0) or 0
                    result[group_key]["vygr_this_year_raw"] = getattr(extra_obj, "vygr_this_year", 0) or 0
                    result[group_key]["vygr_last_year_raw"] = getattr(extra_obj, "vygr_last_year", 0) or 0
                    result[group_key]["pogr_kont_this_year_raw"] = getattr(extra_obj, "pogr_kont_this_year", 0) or 0
                    result[group_key]["pogr_kont_last_year_raw"] = getattr(extra_obj, "pogr_kont_last_year", 0) or 0
                    result[group_key]["vygr_kont_this_year_raw"] = getattr(extra_obj, "vygr_kont_this_year", 0) or 0
                    result[group_key]["vygr_kont_last_year_raw"] = getattr(extra_obj, "vygr_kont_last_year", 0) or 0
                    result[group_key]["income_this_year_raw"] = getattr(extra_obj, "income_this_year", 0) // 1000 or 0
                    result[group_key]["income_last_year_raw"] = getattr(extra_obj, "income_last_year", 0) //  1000 or 0

                result[group_key]["pogr_plan"] += round((getattr(extra_obj, "pogr_plan", 0) or 0) / days_in_month * selected_days)
                result[group_key]["vygr_plan"] += round((getattr(extra_obj, "vygr_plan", 0) or 0) / days_in_month * selected_days)
                result[group_key]["pogr_kont_plan"] += round((getattr(extra_obj, "pogr_kont_plan", 0) or 0) / days_in_month * selected_days)
                result[group_key]["vygr_kont_plan"] += round((getattr(extra_obj, "vygr_kont_plan", 0) or 0) / days_in_month * selected_days)
                result[group_key]["income_plan"] += round((getattr(extra_obj, "income_plan", 0) or 0) / days_in_month * selected_days) // 1000

                result[group_key]["pogr_this_year"] += round((getattr(extra_obj, "pogr_this_year", 0) or 0) / days_in_month * selected_days)
                result[group_key]["pogr_last_year"] += round((getattr(extra_obj, "pogr_last_year", 0) or 0) / days_in_month * selected_days)
                result[group_key]["vygr_this_year"] += round((getattr(extra_obj, "vygr_this_year", 0) or 0) / days_in_month * selected_days)
                result[group_key]["vygr_last_year"] += round((getattr(extra_obj, "vygr_last_year", 0) or 0) / days_in_month * selected_days)
                result[group_key]["pogr_kont_this_year"] += round((getattr(extra_obj, "pogr_kont_this_year", 0) or 0) / days_in_month * selected_days)
                result[group_key]["pogr_kont_last_year"] += round((getattr(extra_obj, "pogr_kont_last_year", 0) or 0) / days_in_month * selected_days)
                result[group_key]["vygr_kont_this_year"] += round((getattr(extra_obj, "vygr_kont_this_year", 0) or 0) / days_in_month * selected_days)
                result[group_key]["vygr_kont_last_year"] += round((getattr(extra_obj, "vygr_kont_last_year", 0) or 0) / days_in_month * selected_days)
                result[group_key]["income_this_year"] += round((getattr(extra_obj, "income_this_year", 0) or 0) / days_in_month * selected_days)
                result[group_key]["income_last_year"] += round((getattr(extra_obj, "income_last_year", 0) or 0) / days_in_month * selected_days) // 1000

                result[group_key]["pogr_diff"] = result[group_key]["pogr_this_year"] - result[group_key]["pogr_last_year"]
                result[group_key]["vygr_diff"] = result[group_key]["vygr_this_year"] - result[group_key]["vygr_last_year"]
                result[group_key]["pogr_kont_diff"] = result[group_key]["pogr_kont_this_year"] - result[group_key]["pogr_kont_last_year"]
                result[group_key]["vygr_kont_diff"] = result[group_key]["vygr_kont_this_year"] - result[group_key]["vygr_kont_last_year"]
                result[group_key]["income_diff"] = (result[group_key]["income_this_year"] - result[group_key]["income_last_year"]) // 1000

        return result

    veshoz_data = _sum_veshoz_between()

    stations_qs = StationProfile.objects.all().order_by("station_name")
    stations_by_name = {_normalize_station_name(x.station_name): x for x in stations_qs}

    groups = []
    grand_total = _make_zero_totals("ИТОГО")
    known_station_names = set()

    for idx, cfg in enumerate(DISPLAY_GROUPS, start=1):
        group_rows = []
        subtotal = _make_zero_totals("ИТОГО")

        for station_name in cfg["stations"]:
            known_station_names.add(_normalize_station_name(station_name))
            station = stations_by_name.get(_normalize_station_name(station_name))

            if station:
                row = _row_to_range_dict(
                    station=station,
                    fact_this=current_data.get(station.id),
                    fact_last=last_year_data.get(station.id),
                    plan_data=scaled_plans.get(station.id),
                )
            else:
                row = _make_empty_range_row(station_name)

            row["is_other"] = False
            row["is_veshoz"] = False
            row["group_key"] = cfg["title"]
            row["is_editable"] = bool(single_full_month and row.get("station_id"))

            group_rows.append(row)
            _add_to_totals(subtotal, row)
            _add_to_totals(grand_total, row)

        if cfg.get("has_veshoz"):
            extra_row = veshoz_data.get(cfg["title"])
            if not extra_row:
                extra_row = _make_empty_range_row("Boshqa Stansiya")
                extra_row.update({
                    "is_other": False,
                    "is_veshoz": True,
                    "group_key": cfg["title"],
                    "is_editable": bool(single_full_month),
                })
            group_rows.append(extra_row)
            _add_to_totals(subtotal, extra_row)
            _add_to_totals(grand_total, extra_row)

        groups.append({
            "index": idx,
            "title": cfg["title"],
            "rows": group_rows,
            "subtotal": subtotal,
        })

    unmatched_rows = []
    unmatched_total = _make_zero_totals("ИТОГО")

    for station_name, station in stations_by_name.items():
        if station_name not in known_station_names:
            row = _row_to_range_dict(
                station=station,
                fact_this=current_data.get(station.id),
                fact_last=last_year_data.get(station.id),
                plan_data=scaled_plans.get(station.id),
            )
            row["is_other"] = True
            row["is_veshoz"] = False
            row["group_key"] = None
            row["is_editable"] = bool(single_full_month and row.get("station_id"))

            unmatched_rows.append(row)
            _add_to_totals(unmatched_total, row)
            _add_to_totals(grand_total, row)

    unmatched_rows.sort(key=lambda x: x["station_name"].lower())

    if unmatched_rows:
        groups.append({
            "index": len(groups) + 1,
            "title": "Прочие станции",
            "rows": unmatched_rows,
            "subtotal": unmatched_total,
        })

    return {
        "from_date": from_date,
        "to_date": to_date,
        "prev_from_date": prev_from_date,
        "prev_to_date": prev_to_date,
        "groups": groups,
        "grand_total": grand_total,
        "all_full_months": all_full_months,
        "single_full_month": single_full_month,
        "month_info": month_info,
        "days_count": (to_date - to_date).days + 1 if False else (to_date - from_date).days + 1,
    }


@transaction.atomic
def kvartalniy_range(request):
    if not request.user.is_superuser:
        return redirect("station_table_1_list")

    if request.method == "POST":
        from_date_str = request.POST.get("from_date")
        to_date_str = request.POST.get("to_date")
    else:
        from_date_str = request.GET.get("from_date")
        to_date_str = request.GET.get("to_date")

    today = timezone.localdate()
    default_from = today.replace(day=1)
    default_to = today

    from_date = _safe_date(from_date_str, default_from)
    to_date = _safe_date(to_date_str, default_to)

    context = _build_kvartalniy_range_context(from_date, to_date)
    return render(request, "kvartalniy_range.html", context)


def kvartalniy_range_export_excel(request):
    if not request.user.is_superuser:
        return redirect("station_table_1_list")

    from_date_str = request.GET.get("from_date")
    to_date_str = request.GET.get("to_date")

    today = timezone.localdate()
    default_from = today.replace(day=1)
    default_to = today

    from_date = _safe_date(from_date_str, default_from)
    to_date = _safe_date(to_date_str, default_to)

    context = _build_kvartalniy_range_context(from_date, to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kvartalniy Range"
    ws.freeze_panes = "B4"

    # ===== styles =====
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    font_normal = Font(name="Times New Roman", size=11, bold=False)
    font_bold = Font(name="Times New Roman", size=11, bold=True)
    font_title = Font(name="Times New Roman", size=12, bold=True)

    fill_title = PatternFill("solid", fgColor="D9D9D9")
    fill_group_header = PatternFill("solid", fgColor="D9D9D9")
    fill_sub_header = PatternFill("solid", fgColor="E7E7E7")
    fill_total = PatternFill("solid", fgColor="F2F2F2")

    green_font = Font(name="Times New Roman", size=11, bold=False, color="008000")
    blue_font = Font(name="Times New Roman", size=11, bold=False, color="0000FF")
    red_font = Font(name="Times New Roman", size=11, bold=False, color="FF0000")

    green_bold_font = Font(name="Times New Roman", size=11, bold=True, color="008000")
    blue_bold_font = Font(name="Times New Roman", size=11, bold=True, color="0000FF")
    red_bold_font = Font(name="Times New Roman", size=11, bold=True, color="FF0000")

    purple_font = Font(name="Times New Roman", size=11, bold=True, color="800080")
    brown_font = Font(name="Times New Roman", size=11, bold=True, color="7C4A03")

    def safe_percent(current, previous):
        current = current or 0
        previous = previous or 0

        if previous == 0:
            return 0 if current == 0 else 100

        return round(((current - previous) / previous) * 100)

    def income_thousand(value):
        """
        Daromad values are stored in full so'm.
        Excel export removes the last 3 digits.

        Examples:
        1234567  -> 1234
        1234999  -> 1234
        999      -> 0
        -1234567 -> -1234
        """
        value = value or 0
        return int(value / 1000)

    def diff_font(value, bold=False):
        if (value or 0) < 0:
            return red_bold_font if bold else red_font

        return font_bold if bold else font_normal

    def percent_font(value, bold=False):
        if (value or 0) < 0:
            return red_bold_font if bold else red_font

        return font_bold if bold else font_normal

    def set_cell(row, col, value, font=None, fill=None, border=None, alignment=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font or font_normal
        c.fill = fill or PatternFill(fill_type=None)
        c.border = border or border_thin
        c.alignment = alignment or center
        return c

    # ===== widths =====
    widths = {
        "A": 22,
        "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
        "G": 10, "H": 10, "I": 10, "J": 10, "K": 10,
        "L": 10, "M": 10, "N": 10, "O": 10, "P": 10,
        "Q": 10, "R": 10, "S": 10, "T": 10,
        "U": 12, "V": 12, "W": 12, "X": 10,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    # ===== title =====
    title = (
        "\"O'ztemiryo'lkonteyner\" AJ ga qarashli Logistika Markazlari va sektorlarida "
        "vagon va konteynerlarni ortib tushirish ishlari va daromad tushumlari to'g'risida tezkor ma'lumotlar\n"
        f"{from_date.strftime('%d.%m.%Y')} — {to_date.strftime('%d.%m.%Y')}  "
        f"taqqoslash: {context['prev_from_date'].strftime('%d.%m.%Y')} — {context['prev_to_date'].strftime('%d.%m.%Y')}"
    )

    ws.merge_cells("A1:X1")
    c = ws["A1"]
    c.value = title
    c.font = font_title
    c.alignment = center
    c.fill = fill_title
    c.border = border_medium

    # ===== headers =====
    ws.merge_cells("A2:A3")
    set_cell(
        2,
        1,
        "LM nomlari",
        font=font_bold,
        fill=fill_group_header,
        border=border_thin,
        alignment=center,
    )
    ws["A3"].border = border_thin

    headers_merged = [
        ("B2:F2", "Ortish vagonda (dona)"),
        ("G2:K2", "Tushirish vagonda (dona)"),
        ("L2:P2", "Ortish konteyner (dona)"),
        ("Q2:T2", "Tushirish konteyner (dona)"),
        ("U2:X2", "Daromad (ming so'm)"),
    ]

    for rng, label in headers_merged:
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = label
        cell.font = font_bold
        cell.alignment = center
        cell.fill = fill_group_header
        cell.border = border_thin

    subheaders = [
        "Reja", "Joriy", "Oldingi", "Farq", "%",
        "Reja", "Joriy", "Oldingi", "Farq", "%",
        "Reja", "Joriy", "Oldingi", "Farq", "%",
        "Joriy", "Oldingi", "Farq", "%",
        "Joriy", "Oldingi", "Farq", "%",
    ]

    for col_idx, label in enumerate(subheaders, start=2):
        set_cell(
            3,
            col_idx,
            label,
            font=font_bold,
            fill=fill_sub_header,
            border=border_thin,
            alignment=center,
        )

    row_num = 4

    # ===== data rows, no group title rows =====
    for group in context["groups"]:
        for row in group["rows"]:
            name_font = purple_font

            if row.get("is_other"):
                name_font = blue_bold_font
            elif row.get("is_veshoz"):
                name_font = brown_font

            pogr_percent = safe_percent(row["pogr_this_year"], row["pogr_last_year"])
            vygr_percent = safe_percent(row["vygr_this_year"], row["vygr_last_year"])
            pogr_kont_percent = safe_percent(row["pogr_kont_this_year"], row["pogr_kont_last_year"])
            vygr_kont_percent = safe_percent(row["vygr_kont_this_year"], row["vygr_kont_last_year"])
            income_percent = safe_percent(row["income_this_year"], row["income_last_year"])

            values = [
                row["station_name"],

                row["pogr_plan"],
                row["pogr_this_year"],
                row["pogr_last_year"],
                row["pogr_diff"],
                f"{pogr_percent}%",

                row["vygr_plan"],
                row["vygr_this_year"],
                row["vygr_last_year"],
                row["vygr_diff"],
                f"{vygr_percent}%",

                row["pogr_kont_plan"],
                row["pogr_kont_this_year"],
                row["pogr_kont_last_year"],
                row["pogr_kont_diff"],
                f"{pogr_kont_percent}%",

                row["vygr_kont_this_year"],
                row["vygr_kont_last_year"],
                row["vygr_kont_diff"],
                f"{vygr_kont_percent}%",

                income_thousand(row["income_this_year"]),
                income_thousand(row["income_last_year"]),
                income_thousand(row["income_diff"]),
                f"{income_percent}%",
            ]

            for col_idx, value in enumerate(values, start=1):
                font = font_normal
                align = center

                if col_idx == 1:
                    font = name_font
                    align = left

                elif col_idx in (3, 8, 13, 17, 21):
                    font = green_font

                elif col_idx in (4, 9, 14, 18, 22):
                    font = blue_font

                elif col_idx in (5, 10, 15, 19, 23):
                    diff_val = None

                    if col_idx == 5:
                        diff_val = row["pogr_diff"]
                    elif col_idx == 10:
                        diff_val = row["vygr_diff"]
                    elif col_idx == 15:
                        diff_val = row["pogr_kont_diff"]
                    elif col_idx == 19:
                        diff_val = row["vygr_kont_diff"]
                    elif col_idx == 23:
                        diff_val = row["income_diff"]

                    font = diff_font(diff_val, bold=False)

                elif col_idx in (6, 11, 16, 20, 24):
                    percent_val = None

                    if col_idx == 6:
                        percent_val = pogr_percent
                    elif col_idx == 11:
                        percent_val = vygr_percent
                    elif col_idx == 16:
                        percent_val = pogr_kont_percent
                    elif col_idx == 20:
                        percent_val = vygr_kont_percent
                    elif col_idx == 24:
                        percent_val = income_percent

                    font = percent_font(percent_val, bold=False)

                set_cell(
                    row_num,
                    col_idx,
                    value,
                    font=font,
                    border=border_thin,
                    alignment=align,
                )

            row_num += 1

        # ===== subtotal row =====
        subtotal = group["subtotal"]

        pogr_percent = safe_percent(subtotal["pogr_this_year"], subtotal["pogr_last_year"])
        vygr_percent = safe_percent(subtotal["vygr_this_year"], subtotal["vygr_last_year"])
        pogr_kont_percent = safe_percent(subtotal["pogr_kont_this_year"], subtotal["pogr_kont_last_year"])
        vygr_kont_percent = safe_percent(subtotal["vygr_kont_this_year"], subtotal["vygr_kont_last_year"])
        income_percent = safe_percent(subtotal["income_this_year"], subtotal["income_last_year"])

        subtotal_values = [
            subtotal["station_name"],

            subtotal["pogr_plan"],
            subtotal["pogr_this_year"],
            subtotal["pogr_last_year"],
            subtotal["pogr_diff"],
            f"{pogr_percent}%",

            subtotal["vygr_plan"],
            subtotal["vygr_this_year"],
            subtotal["vygr_last_year"],
            subtotal["vygr_diff"],
            f"{vygr_percent}%",

            subtotal["pogr_kont_plan"],
            subtotal["pogr_kont_this_year"],
            subtotal["pogr_kont_last_year"],
            subtotal["pogr_kont_diff"],
            f"{pogr_kont_percent}%",

            subtotal["vygr_kont_this_year"],
            subtotal["vygr_kont_last_year"],
            subtotal["vygr_kont_diff"],
            f"{vygr_kont_percent}%",

            income_thousand(subtotal["income_this_year"]),
            income_thousand(subtotal["income_last_year"]),
            income_thousand(subtotal["income_diff"]),
            f"{income_percent}%",
        ]

        for col_idx, value in enumerate(subtotal_values, start=1):
            font = font_bold
            align = center

            if col_idx == 1:
                align = left

            elif col_idx in (4, 9, 14, 18, 22):
                font = green_bold_font

            elif col_idx in (5, 10, 15, 19, 23):
                diff_val = None

                if col_idx == 5:
                    diff_val = subtotal["pogr_diff"]
                elif col_idx == 10:
                    diff_val = subtotal["vygr_diff"]
                elif col_idx == 15:
                    diff_val = subtotal["pogr_kont_diff"]
                elif col_idx == 19:
                    diff_val = subtotal["vygr_kont_diff"]
                elif col_idx == 23:
                    diff_val = subtotal["income_diff"]

                font = diff_font(diff_val, bold=True)

            elif col_idx in (6, 11, 16, 20, 24):
                percent_val = None

                if col_idx == 6:
                    percent_val = pogr_percent
                elif col_idx == 11:
                    percent_val = vygr_percent
                elif col_idx == 16:
                    percent_val = pogr_kont_percent
                elif col_idx == 20:
                    percent_val = vygr_kont_percent
                elif col_idx == 24:
                    percent_val = income_percent

                font = percent_font(percent_val, bold=True)

            set_cell(
                row_num,
                col_idx,
                value,
                font=font,
                fill=fill_total,
                border=border_thin,
                alignment=align,
            )

        row_num += 1

    # ===== grand total =====
    grand = context["grand_total"]

    pogr_percent = safe_percent(grand["pogr_this_year"], grand["pogr_last_year"])
    vygr_percent = safe_percent(grand["vygr_this_year"], grand["vygr_last_year"])
    pogr_kont_percent = safe_percent(grand["pogr_kont_this_year"], grand["pogr_kont_last_year"])
    vygr_kont_percent = safe_percent(grand["vygr_kont_this_year"], grand["vygr_kont_last_year"])
    income_percent = safe_percent(grand["income_this_year"], grand["income_last_year"])

    grand_values = [
        "Всего",

        grand["pogr_plan"],
        grand["pogr_this_year"],
        grand["pogr_last_year"],
        grand["pogr_diff"],
        f"{pogr_percent}%",

        grand["vygr_plan"],
        grand["vygr_this_year"],
        grand["vygr_last_year"],
        grand["vygr_diff"],
        f"{vygr_percent}%",

        grand["pogr_kont_plan"],
        grand["pogr_kont_this_year"],
        grand["pogr_kont_last_year"],
        grand["pogr_kont_diff"],
        f"{pogr_kont_percent}%",

        grand["vygr_kont_this_year"],
        grand["vygr_kont_last_year"],
        grand["vygr_kont_diff"],
        f"{vygr_kont_percent}%",

        income_thousand(grand["income_this_year"]),
        income_thousand(grand["income_last_year"]),
        income_thousand(grand["income_diff"]),
        f"{income_percent}%",
    ]

    for col_idx, value in enumerate(grand_values, start=1):
        font = font_bold
        align = center

        if col_idx == 1:
            align = left

        elif col_idx in (4, 9, 14, 18, 22):
            font = green_bold_font

        elif col_idx in (5, 10, 15, 19, 23):
            diff_val = None

            if col_idx == 5:
                diff_val = grand["pogr_diff"]
            elif col_idx == 10:
                diff_val = grand["vygr_diff"]
            elif col_idx == 15:
                diff_val = grand["pogr_kont_diff"]
            elif col_idx == 19:
                diff_val = grand["vygr_kont_diff"]
            elif col_idx == 23:
                diff_val = grand["income_diff"]

            font = diff_font(diff_val, bold=True)

        elif col_idx in (6, 11, 16, 20, 24):
            percent_val = None

            if col_idx == 6:
                percent_val = pogr_percent
            elif col_idx == 11:
                percent_val = vygr_percent
            elif col_idx == 16:
                percent_val = pogr_kont_percent
            elif col_idx == 20:
                percent_val = vygr_kont_percent
            elif col_idx == 24:
                percent_val = income_percent

            font = percent_font(percent_val, bold=True)

        set_cell(
            row_num,
            col_idx,
            value,
            font=font,
            fill=fill_total,
            border=border_thin,
            alignment=align,
        )

    # ===== medium border around title =====
    for col in range(1, 25):
        ws.cell(1, col).border = Border(
            left=medium if col == 1 else thin,
            right=medium if col == 24 else thin,
            top=medium,
            bottom=medium,
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"kvartalniy_range_{from_date}_{to_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response

def _safe_pdf_filename_part(value: str, default: str) -> str:
    value = str(value or "").strip()

    if not value:
        return default

    value = re.sub(r"[^0-9A-Za-z_\-\.]", "_", value)
    return value[:40] or default


def kvartalniy_range_export_pdf(request):
    """
    Kvartalniy umumiy/range hisobotini PDF qilib chiqaradi.

    Bu variant:
    - 1 sahifada qoladi;
    - eski yaxshi ko'rinishni saqlaydi;
    - jadvalni faqat vizual tarzda pastga cho'zadi;
    - row height/padding bilan o'ynamaydi, shuning uchun 2-sahifaga o'tmaydi.
    """

    query_string = request.GET.urlencode()

    report_url = request.build_absolute_uri(
        reverse("kvartalniy_um") + (f"?{query_string}" if query_string else "")
    )

    from_date = _safe_pdf_filename_part(request.GET.get("from_date"), "from")
    to_date = _safe_pdf_filename_part(request.GET.get("to_date"), "to")
    filename = f"kvartalniy_{from_date}_{to_date}.pdf"

    pdf_css = """
      @page {
        size: A4 landscape;
        margin: 2mm;
      }

      html,
      body {
        margin: 0 !important;
        padding: 0 !important;
        width: 1900px !important;
        min-width: 1900px !important;
        max-width: 1900px !important;
        background: #ffffff !important;
        overflow: visible !important;
      }

      * {
        -webkit-print-color-adjust: exact !important;
        print-color-adjust: exact !important;
        box-shadow: none !important;
        text-shadow: none !important;
      }

      header,
      nav,
      aside,
      footer,
      .sidebar,
      .navbar,
      .topbar,
      .main-sidebar,
      .app-sidebar,
      .layout-sidebar,
      .kv-mini-nav,
      .kv-title-wrap,
      .kv-actions,
      .kv-status,
      .kv-note,
      .messages,
      .alert,
      button,
      .kv-btn,
      input,
      label {
        display: none !important;
      }

      form {
        display: block !important;
        visibility: visible !important;
        opacity: 1 !important;
        margin: 0 !important;
        padding: 0 !important;
        width: 1900px !important;
        max-width: none !important;
        min-width: 1900px !important;
        height: auto !important;
        max-height: none !important;
        overflow: visible !important;
        background: #ffffff !important;
      }

      main,
      .content,
      .content-wrapper,
      .page-content,
      .container,
      .container-fluid,
      .kv-page,
      .kv-shell,
      .kv-card,
      .kv-wrap,
      .kv-scroll {
        display: block !important;
        visibility: visible !important;
        opacity: 1 !important;
        margin: 0 !important;
        padding: 0 !important;
        width: 1900px !important;
        max-width: none !important;
        min-width: 1900px !important;
        height: auto !important;
        max-height: none !important;
        overflow: visible !important;
        background: #ffffff !important;
        border-radius: 0 !important;
      }

      .kv-card {
        border: 0 !important;
      }

      .kv-wrap {
        border: 2px solid #111827 !important;
        overflow: visible !important;
      }

      .kv-table {
        display: table !important;
        visibility: visible !important;
        opacity: 1 !important;
        width: 1900px !important;
        min-width: 1900px !important;
        max-width: 1900px !important;
        table-layout: fixed !important;
        border-collapse: collapse !important;
        border-spacing: 0 !important;
        background: #ffffff !important;
        color: #000000 !important;
        font-family: "Times New Roman", Times, serif !important;

        transform: scaleY(2.20) !important;
        transform-origin: top left !important;
      }

      .kv-table thead {
        display: table-header-group !important;
      }

      .kv-table tbody {
        display: table-row-group !important;
      }

      .kv-table tr {
        display: table-row !important;
      }

      .kv-table th,
      .kv-table td {
        display: table-cell !important;
        visibility: visible !important;
        opacity: 1 !important;
        border: 1px solid #3b3b3b !important;
        padding: 1px 2px !important;
        font-size: 6.2px !important;
        line-height: 1.02 !important;
        text-align: center !important;
        vertical-align: middle !important;
        white-space: nowrap !important;
        color: #000000 !important;
        background: #ffffff !important;
        font-weight: 700 !important;
      }

      .kv-table thead th {
        background: #eef2f7 !important;
        color: #000000 !important;
        font-weight: 900 !important;
      }

      .kv-main-title,
      .kv-super-title th {
        font-size: 8px !important;
        line-height: 1.08 !important;
        padding: 3px 4px !important;
        background: #f8fafc !important;
        border-bottom: 2px solid #111827 !important;
        font-weight: 900 !important;
      }

      .name-col,
      .station-col,
      .kv-station-col {
        width: 150px !important;
        min-width: 150px !important;
        max-width: 150px !important;
        text-align: left !important;
        padding-left: 4px !important;
      }

      .regular-col,
      .percent-col,
      .kv-col-normal {
        width: 64px !important;
        min-width: 64px !important;
        max-width: 64px !important;
      }

      .kv-col-mid {
        width: 70px !important;
        min-width: 70px !important;
        max-width: 70px !important;
      }

      .money-col,
      .kv-col-income {
        width: 105px !important;
        min-width: 105px !important;
        max-width: 105px !important;
      }

      .sep-l {
        border-left: 2px solid #111827 !important;
      }

      .sep-r {
        border-right: 2px solid #111827 !important;
      }

      .sep-b {
        border-bottom: 2px solid #111827 !important;
      }

      .group-title-row td,
      .kv-group {
        background: #e9eef5 !important;
        font-weight: 900 !important;
      }

      .subtotal-row td,
      .grand-total-row td {
        background: #fff4f4 !important;
        font-weight: 900 !important;
      }

      .green,
      .readonly-green {
        color: #00945a !important;
      }

      .blue,
      .readonly-blue {
        color: #0047c7 !important;
      }

      .red,
      .readonly-red {
        color: #ff0000 !important;
      }

      .purple {
        color: #4f46e5 !important;
      }

      tr,
      td,
      th {
        break-inside: avoid !important;
        page-break-inside: avoid !important;
      }
    """

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ],
        )

        context = browser.new_context(
            viewport={
                "width": 1900,
                "height": 1200,
            },
            device_scale_factor=1,
        )

        base_url = request.build_absolute_uri("/")

        for name, value in request.COOKIES.items():
            context.add_cookies([
                {
                    "name": name,
                    "value": value,
                    "url": base_url,
                }
            ])

        page = context.new_page()

        page.goto(
            report_url,
            wait_until="domcontentloaded",
            timeout=60000,
        )

        page.wait_for_selector(
            ".kv-table",
            state="attached",
            timeout=60000,
        )

        page.add_style_tag(content=pdf_css)

        page.evaluate(
            """
            () => {
              const table = document.querySelector(".kv-table");

              if (!table) {
                throw new Error("PDF export error: .kv-table not found");
              }

              let el = table;

              while (el && el !== document.body) {
                el.style.display = "";
                el.style.visibility = "visible";
                el.style.opacity = "1";
                el.style.overflow = "visible";
                el.style.maxHeight = "none";
                el.style.height = "auto";
                el = el.parentElement;
              }

              table.style.display = "table";
              table.style.visibility = "visible";
              table.style.opacity = "1";
              table.style.transform = "scaleY(2.20)";
              table.style.transformOrigin = "top left";
            }
            """
        )

        pdf_bytes = page.pdf(
            format="A4",
            landscape=True,
            print_background=True,
            prefer_css_page_size=True,
            scale=0.58,
            margin={
                "top": "2mm",
                "right": "2mm",
                "bottom": "2mm",
                "left": "2mm",
            },
        )

        browser.close()

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response