from django.http import HttpResponse
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from accounts.models import StationProfile
from reports.forms import TABLE1_FIELDS
from reports.models import StationDailyTable1
from reports.views import TERMINAL_NAME_KEY, _apply_itogo_rules, _parse_date, _station_display_name, _terminal_blocks_for_station_date, staff_required

from django.http import HttpResponse
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.http import HttpResponse
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


from django.http import HttpResponse
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


from django.http import HttpResponse
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


DISPLAY_GROUPS = [
    {
        "title": "group1",
        "stations": [
            "Chuqursoy LM",
            "Toshkent LM",
            "Sergeli LM",
            "Ulug'bek LM",
            "Marokand LM",
            "Jaloir LM",
            "Ohangaron LM",
            "Nazarbek LM",
            "Urtavul LM",
            "Sirdaryo LM",
            "Jizzax LM",
            "Ablik LM",
            "To'ytepa",
        ],
        "has_veshoz": True,
    },
    {
        "title": "group2",
        "stations": [
            "Qo'qon LM",
            "Rovuston LM",
            "Marg'ilon LM",
            "Axtachi LM",
            "Asaka LM",
        ],
        "has_veshoz": True,
    },
    {
        "title": "group3",
        "stations": [
            "Buxoro LM",
            "Tinchlik LM",
            "Karmana LM",
            "Yangi-Zarafshon LM",
            
        ],
        "has_veshoz": True,
    },
    {
        "title": "group4",
        "stations": [
            "Qarshi LM",
            "Dehqonobod LM",
        ],
        "has_veshoz": True,
    },
    {
        "title": "group5",
        "stations": [
            "Termiz LM",
        ],
        "has_veshoz": True,
    },
    {
        "title": "group6",
        "stations": [
            "Nukus LM",
            "Kirkkiz LM",
            "Urganch LM",
            "Pitnyak LM",
        ],
        "has_veshoz": True,
    },
]

from django.http import HttpResponse
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


@staff_required
def admin_table1_report_excel_view(request, date_str):
    d = _parse_date(date_str)

    User = get_user_model()
    users = (
        User.objects
        .exclude(is_staff=True)
        .exclude(is_superuser=True)
        .order_by("username")
    )

    FIELDS = [key for key, _label in TABLE1_FIELDS]

    def _to_int(v):
        if v in (None, "", "—", "-", "–"):
            return 0
        if isinstance(v, str):
            v = v.replace("\xa0", "").replace(" ", "").replace(",", "")
        try:
            return int(float(v))
        except Exception:
            return 0

    def _sum_into(dst: dict, src: dict):
        for k in FIELDS:
            dst[k] = dst.get(k, 0) + _to_int(src.get(k, 0))

    def _station_order_index(name):
        for gi, group in enumerate(DISPLAY_GROUPS):
            stations = group.get("stations", [])
            for si, st_name in enumerate(stations):
                if (st_name or "").strip().lower() == (name or "").strip().lower():
                    return (gi, si)
        return (999, (name or "").lower())

    station_list = []

    for u in users:
        try:
            sp = StationProfile.objects.select_related("user").get(user=u)
        except StationProfile.DoesNotExist:
            continue

        has_night = bool(sp.status)

        sent = StationDailyTable1.objects.filter(
            station_user=u,
            date=d,
            submitted_at__isnull=False,
        ).exists()
        if not sent:
            continue

        blocks = _terminal_blocks_for_station_date(u, d, force_new=False)

        terminals = []
        for b in blocks:
            day_obj = StationDailyTable1.objects.filter(
                station_user=u, date=d, shift="day", block=b
            ).first()
            night_obj = StationDailyTable1.objects.filter(
                station_user=u, date=d, shift="night", block=b
            ).first()

            day_raw = (day_obj.data or {}) if day_obj else {}
            night_raw = (night_obj.data or {}) if (night_obj and has_night) else {}

            day_data = _apply_itogo_rules(day_raw)
            night_data = _apply_itogo_rules(night_raw) if has_night else {}

            total_data = {k: 0 for k in FIELDS}
            _sum_into(total_data, day_data)
            if has_night:
                _sum_into(total_data, night_data)
            total_data = _apply_itogo_rules(total_data)

            term_name = (
                (day_raw or {}).get(TERMINAL_NAME_KEY)
                or (night_raw or {}).get(TERMINAL_NAME_KEY)
                or ""
            )

            terminals.append({
                "block": b,
                "terminal_name": term_name,
                "day_data": day_data,
                "night_data": night_data,
                "total_data": total_data,
            })

        sum_total = {k: 0 for k in FIELDS}
        for t in terminals:
            _sum_into(sum_total, t["day_data"])
            if has_night:
                _sum_into(sum_total, t["night_data"])
        sum_total = _apply_itogo_rules(sum_total)

        station_name = _station_display_name(u)

        station_list.append({
            "name": station_name,
            "login": getattr(u, "username", ""),
            "user_id": u.id,
            "status": has_night,
            "terminals": terminals,
            "sum_total": sum_total,
            "_order": _station_order_index(station_name),
        })

    station_list.sort(key=lambda x: x["_order"])

    grand_total = {k: 0 for k in FIELDS}
    for st in station_list:
        _sum_into(grand_total, st.get("sum_total") or {})
    grand_total = _apply_itogo_rules(grand_total)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Table1_{d.strftime('%d_%m_%Y')}"

    thin = Side(style="thin", color="B7BDC7")
    medium = Side(style="medium", color="7C8591")
    thick = Side(style="thick", color="59606A")

    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)

    font_title = Font(name="Arial", size=13, bold=True)
    font_header = Font(name="Arial", size=9, bold=True)
    font_body = Font(name="Arial", size=9)
    font_total = Font(name="Arial", size=9, bold=True)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    fill_header = PatternFill("solid", fgColor="E7E7E7")
    fill_pink = PatternFill("solid", fgColor="EEDFE3")
    fill_green = PatternFill("solid", fgColor="DDEAD7")
    fill_green_dark = PatternFill("solid", fgColor="C8DEBF")
    fill_yellow = PatternFill("solid", fgColor="F5EFC8")
    fill_blue = PatternFill("solid", fgColor="E4EFF3")
    fill_blue_dark = PatternFill("solid", fgColor="C7E4EA")
    fill_total = PatternFill("solid", fgColor="F3EBCB")
    fill_white = PatternFill("solid", fgColor="FFFFFF")

    columns = [
        ("station_name", "LM nomi", 18),
        ("shift", "Smena", 9),
        ("terminal_name", "Terminal", 15),

        ("podano_lc", "LMga berildi", 9),
        ("k_podache_so_st", "St’dan berishga", 9),

        ("vygr_ft", "ft", 6),
        ("vygr_cont", "kont", 6),
        ("vygr_kr", "kr", 6),
        ("vygr_pv", "pv", 6),
        ("vygr_proch", "boshqa", 7),
        ("vygr_itogo", "jami", 6),
        ("vygr_itogo_kon", "jami kont", 7),

        ("pod_vygr_ft", "ft", 6),
        ("pod_vygr_cont", "kont", 6),
        ("pod_vygr_kr", "kr", 6),
        ("pod_vygr_pv", "pv", 6),
        ("pod_vygr_proch", "boshqa", 7),
        ("pod_vygr_itogo", "jami", 6),
        ("pod_vygr_itogo_kon", "jami kont", 7),

        ("uborka", "Yig‘ishtirish", 7),

        ("pogr_ft", "ft", 6),
        ("pogr_cont", "kont", 6),
        ("pogr_kr", "kr", 6),
        ("pogr_pv", "pv", 6),
        ("pogr_proch", "boshqa", 7),
        ("pogr_itogo", "jami", 6),
        ("pogr_itogo_kon", "jami kont", 7),

        ("pod_pogr_ft", "ft", 6),
        ("pod_pogr_cont", "kont", 6),
        ("pod_pogr_kr", "kr", 6),
        ("pod_pogr_pv", "pv", 6),
        ("pod_pogr_proch", "boshqa", 7),
        ("pod_pogr_itogo", "jami", 6),
        ("pod_pogr_itogo_kon", "jami kont", 7),

        ("income_daily", "sutkalik daromad", 12),
    ]

    for idx, (_, _, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    def cell(row, col, value="", font=None, fill=None, border=None, align=None, is_number=False):
        c = ws.cell(row=row, column=col, value=value)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        if border:
            c.border = border
        if align:
            c.alignment = align
        if is_number:
            c.number_format = '#,##0'
        return c

    def set_range_thick_columns_only(r1, c1, r2, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cur = ws.cell(r, c).border
                left = cur.left
                right = cur.right
                top = cur.top
                bottom = cur.bottom

                if c == c1:
                    left = thick
                if c == c2:
                    right = thick

                ws.cell(r, c).border = Border(
                    left=left,
                    right=right,
                    top=top,
                    bottom=bottom,
                )

    def fill_for_key(key, is_total=False):
        if is_total:
            return fill_total
        if key in ("podano_lc", "k_podache_so_st"):
            return fill_pink
        if key.startswith("vygr_"):
            return fill_green
        if key.startswith("pod_vygr_"):
            return fill_green_dark
        if key == "uborka":
            return fill_yellow
        if key.startswith("pogr_"):
            return fill_blue
        if key.startswith("pod_pogr_"):
            return fill_blue_dark
        if key == "income_daily":
            return fill_white
        return fill_white

    row = 1
    last_col = len(columns)

    title_text = f"\"O‘ztemiryo‘lkonteyner\" AJ logistika markazlari bo‘yicha sutkalik operativ ma’lumot — {d.strftime('%d.%m.%Y')}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell(row, 1, title_text, font=font_title, align=align_center)
    row += 2

    header_row_1 = row
    header_row_2 = row + 1

    ws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_2, end_column=1)
    ws.merge_cells(start_row=header_row_1, start_column=2, end_row=header_row_2, end_column=2)
    ws.merge_cells(start_row=header_row_1, start_column=3, end_row=header_row_2, end_column=3)

    ws.merge_cells(start_row=header_row_1, start_column=4, end_row=header_row_2, end_column=4)
    ws.merge_cells(start_row=header_row_1, start_column=5, end_row=header_row_2, end_column=5)

    ws.merge_cells(start_row=header_row_1, start_column=6, end_row=header_row_1, end_column=12)
    ws.merge_cells(start_row=header_row_1, start_column=13, end_row=header_row_1, end_column=19)
    ws.merge_cells(start_row=header_row_1, start_column=20, end_row=header_row_2, end_column=20)
    ws.merge_cells(start_row=header_row_1, start_column=21, end_row=header_row_1, end_column=27)
    ws.merge_cells(start_row=header_row_1, start_column=28, end_row=header_row_1, end_column=34)
    ws.merge_cells(start_row=header_row_1, start_column=35, end_row=header_row_2, end_column=35)

    cell(header_row_1, 1, "LM nomi", font=font_header, fill=fill_header, border=border_medium, align=align_center)
    cell(header_row_1, 2, "Smena", font=font_header, fill=fill_header, border=border_medium, align=align_center)
    cell(header_row_1, 3, "Terminal", font=font_header, fill=fill_header, border=border_medium, align=align_center)

    cell(header_row_1, 4, "LMga berildi", font=font_header, fill=fill_pink, border=border_medium, align=align_center)
    cell(header_row_1, 5, "St’dan berishga", font=font_header, fill=fill_pink, border=border_medium, align=align_center)

    cell(header_row_1, 6, "Tushirish", font=font_header, fill=fill_header, border=border_medium, align=align_center)
    cell(header_row_1, 13, "Tushirishda", font=font_header, fill=fill_header, border=border_medium, align=align_center)
    cell(header_row_1, 20, "Yig‘ishtirish", font=font_header, fill=fill_yellow, border=border_medium, align=align_center)
    cell(header_row_1, 21, "Ortish", font=font_header, fill=fill_header, border=border_medium, align=align_center)
    cell(header_row_1, 28, "Ortishda", font=font_header, fill=fill_header, border=border_medium, align=align_center)
    cell(header_row_1, 35, "sutkalik daromad", font=font_header, fill=fill_header, border=border_medium, align=align_center)

    subheaders = [
        (6,  "ft",        fill_header),
        (7,  "kont",      fill_header),
        (8,  "kr",        fill_header),
        (9,  "pv",        fill_header),
        (10, "boshqa",    fill_header),
        (11, "jami",      fill_header),
        (12, "jami kont", fill_header),

        (13, "ft",        fill_header),
        (14, "kont",      fill_header),
        (15, "kr",        fill_header),
        (16, "pv",        fill_header),
        (17, "boshqa",    fill_header),
        (18, "jami",      fill_header),
        (19, "jami kont", fill_header),

        (21, "ft",        fill_header),
        (22, "kont",      fill_header),
        (23, "kr",        fill_header),
        (24, "pv",        fill_header),
        (25, "boshqa",    fill_header),
        (26, "jami",      fill_header),
        (27, "jami kont", fill_header),

        (28, "ft",        fill_header),
        (29, "kont",      fill_header),
        (30, "kr",        fill_header),
        (31, "pv",        fill_header),
        (32, "boshqa",    fill_header),
        (33, "jami",      fill_header),
        (34, "jami kont", fill_header),
    ]

    for col_num, title, fill in subheaders:
        cell(header_row_2, col_num, title, font=font_header, fill=fill, border=border_medium, align=align_center)

    for r in range(header_row_1, header_row_2 + 1):
        for c in range(1, last_col + 1):
            ws.cell(r, c).border = border_medium
            if ws.cell(r, c).alignment is None:
                ws.cell(r, c).alignment = align_center

    set_range_thick_columns_only(header_row_1, 6, header_row_2, 12)
    set_range_thick_columns_only(header_row_1, 13, header_row_2, 19)
    set_range_thick_columns_only(header_row_1, 21, header_row_2, 27)
    set_range_thick_columns_only(header_row_1, 28, header_row_2, 34)
    set_range_thick_columns_only(header_row_1, 35, header_row_2, 35)

    row = header_row_2 + 1
    terminal_field_keys = [c[0] for c in columns[3:]]

    def write_terminal_data_row(row_num, terminal_name, source_data, is_total=False):
        cell(
            row_num, 3, terminal_name or "-",
            font=font_body if not is_total else font_total,
            fill=fill_total if is_total else fill_white,
            border=border_thin if not is_total else border_medium,
            align=align_left
        )

        for col_idx, key in enumerate(terminal_field_keys, start=4):
            num_value = _to_int(source_data.get(key, 0))
            cell(
                row_num,
                col_idx,
                num_value,
                font=font_total if is_total else font_body,
                fill=fill_for_key(key, is_total=is_total),
                border=border_thin if not is_total else border_medium,
                align=align_center,
                is_number=True
            )

    if not station_list:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
        cell(row, 1, "Нет данных по этой дате", font=font_body, border=border_thin, align=align_center)
        row += 1
    else:
        for st in station_list:
            terms = st["terminals"]
            has_night = bool(st["status"])

            night_rows = len(terms) if has_night else 0
            day_rows = len(terms)
            show_total_row = has_night
            total_rows = 1 if show_total_row else 0
            station_rows_total = night_rows + day_rows + total_rows

            station_start_row = row

            ws.merge_cells(
                start_row=station_start_row,
                start_column=1,
                end_row=station_start_row + station_rows_total - 1,
                end_column=1
            )
            cell(station_start_row, 1, st["name"], font=font_total, border=border_thin, align=align_left)

            if has_night and night_rows:
                ws.merge_cells(start_row=row, start_column=2, end_row=row + night_rows - 1, end_column=2)
                cell(row, 2, "tun", font=font_total, border=border_thin, align=align_center)

                for t in terms:
                    write_terminal_data_row(
                        row_num=row,
                        terminal_name=t["terminal_name"],
                        source_data=t["night_data"],
                        is_total=False
                    )
                    row += 1

            if day_rows:
                ws.merge_cells(start_row=row, start_column=2, end_row=row + day_rows - 1, end_column=2)
                cell(row, 2, "kun", font=font_total, border=border_thin, align=align_center)

                for t in terms:
                    write_terminal_data_row(
                        row_num=row,
                        terminal_name=t["terminal_name"],
                        source_data=t["day_data"],
                        is_total=False
                    )
                    row += 1

            if show_total_row:
                cell(row, 2, "jami", font=font_total, fill=fill_total, border=border_medium, align=align_center)
                cell(row, 3, "", font=font_total, fill=fill_total, border=border_medium, align=align_center)

                for col_idx, key in enumerate(terminal_field_keys, start=4):
                    num_value = _to_int(st["sum_total"].get(key, 0))
                    cell(
                        row,
                        col_idx,
                        num_value,
                        font=font_total,
                        fill=fill_total,
                        border=border_medium,
                        align=align_center,
                        is_number=True
                    )

                end_outline_row = row
                row += 1
            else:
                end_outline_row = row - 1

            set_range_thick_columns_only(station_start_row, 6, end_outline_row, 12)
            set_range_thick_columns_only(station_start_row, 13, end_outline_row, 19)
            set_range_thick_columns_only(station_start_row, 21, end_outline_row, 27)
            set_range_thick_columns_only(station_start_row, 28, end_outline_row, 34)
            set_range_thick_columns_only(station_start_row, 35, end_outline_row, 35)

        cell(row, 1, "Umumiy", font=font_total, fill=fill_total, border=border_medium, align=align_left)
        cell(row, 2, "", font=font_total, fill=fill_total, border=border_medium, align=align_center)
        cell(row, 3, "", font=font_total, fill=fill_total, border=border_medium, align=align_center)

        for col_idx, key in enumerate(terminal_field_keys, start=4):
            num_value = _to_int(grand_total.get(key, 0))
            cell(
                row,
                col_idx,
                num_value,
                font=font_total,
                fill=fill_total,
                border=border_medium,
                align=align_center,
                is_number=True
            )

        set_range_thick_columns_only(row, 6, row, 12)
        set_range_thick_columns_only(row, 13, row, 19)
        set_range_thick_columns_only(row, 21, row, 27)
        set_range_thick_columns_only(row, 28, row, 34)
        set_range_thick_columns_only(row, 35, row, 35)
        row += 1

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[header_row_1].height = 24
    ws.row_dimensions[header_row_2].height = 52

    for r in range(header_row_2 + 1, row):
        if ws.cell(r, 2).value == "jami" or ws.cell(r, 1).value == "Umumiy":
            ws.row_dimensions[r].height = 20
        else:
            ws.row_dimensions[r].height = 19

    ws.freeze_panes = "D4"

    filename = f'admin_table1_{d.strftime("%Y_%m_%d")}.xlsx'
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response