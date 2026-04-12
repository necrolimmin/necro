from datetime import date as dt_date, datetime
from functools import wraps
import io
import json

from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.core.exceptions import FieldError
from django.core.paginator import Paginator
from django.db.models import Sum, Max
from django.http import HttpResponseNotAllowed, HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.urls import reverse
from django.views.decorators.http import require_GET, require_POST

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from accounts.models import StationProfile
from .models import StationDailyTable1, StationDailyTable2, KPIValue, Notification, NotificationRead
from .forms import TABLE1_FIELDS


# =========================
# helpers
# =========================

def staff_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            return redirect("station_table_1_list")
        return view_func(request, *args, **kwargs)
    return wrapper


def _parse_date(date_str: str) -> dt_date:
    return datetime.strptime(date_str, "%Y-%m-%d").date()


def _read_int(raw: str) -> int:
    raw = (raw or "").strip()
    if raw == "":
        return 0
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        return 0


def _read_int_post(request, name: str) -> int:
    return _read_int(request.POST.get(name))


def _int0(v):
    try:
        return int(v or 0)
    except (TypeError, ValueError):
        return 0


def _dget(d: dict, key: str, default=0) -> int:
    try:
        return int((d or {}).get(key) or default)
    except (TypeError, ValueError):
        return int(default)


# -------------------------
# NEW: терминалы = блоки
# -------------------------

TERMINAL_NAME_KEY = "terminal_name"


def _station_has_night(user) -> bool:
    pro = StationProfile.objects.get(user=user)
    return bool(pro.status) or False


def _terminal_blocks_for_station_date(user, d: dt_date, *, force_new: bool = False) -> list[int]:
    if force_new:
        return [1]

    blocks = list(
        StationDailyTable1.objects
        .filter(station_user=user, date=d)
        .values_list("block", flat=True)
        .distinct()
        .order_by("block")
    )
    return blocks or [1]


def _is_table1_submitted(user, d: dt_date):
    if not user:
        return False, None

    last = (
        StationDailyTable1.objects
        .filter(
            station_user=user,
            date=d,
            shift="total",
            submitted_at__isnull=False,
        )
        .aggregate(last=Max("submitted_at"))
        .get("last")
    )

    return (last is not None), last


# =========================
# SPECIAL KEYS: ROW 22 SPLIT (TABLE2)
# =========================

R22_G_TOTAL = "r22g_total"
R22_G_KTK   = "r22g_ktk"
R22_P_TOTAL = "r22p_total"
R22_P_KTK   = "r22p_ktk"


# =========================
# ADMIN REPORTS
# =========================

@staff_required
def admin_report_1(request):
    return redirect("admin_table1_reports")


@staff_required
def admin_report_2(request):
    d_str = request.GET.get("date")
    d = dt_date.today() if not d_str else datetime.strptime(d_str, "%Y-%m-%d").date()

    agg = (
        KPIValue.objects.filter(date=d)
        .values("kpi__code", "kpi__name", "kpi__order")
        .annotate(
            sum_total=Sum("value_total"),
            sum_ktk=Sum("value_ktk"),
            sum_income=Sum("income"),
        )
        .order_by("kpi__order")
    )
    return render(request, "admin_report_2.html", {"date": d, "agg": agg})


# =========================
# TABLE 1 (STATION)
# =========================

@login_required
def station_table_1_list(request):
    if request.user.is_staff or request.user.is_superuser:
        return redirect("admin_table1_reports")

    qs_dates = (
        StationDailyTable1.objects
        .filter(station_user=request.user, shift="total")
        .values("date")
        .annotate(last_submitted_at=Max("submitted_at"))
        .order_by("-date")
    )

    per_page = _read_int(request.GET.get("per_page")) or 10
    if per_page not in (5, 10, 20, 50):
        per_page = 10

    paginator = Paginator(list(qs_dates), per_page)
    page_number = request.GET.get("page") or 1
    page_obj = paginator.get_page(page_number)

    rows = [{
        "date": r["date"],
        "year": r["date"].year,
        "submitted_at": r["last_submitted_at"],
    } for r in page_obj.object_list]

    existing_dates = set(r["date"] for r in qs_dates)

    return render(request, "station_table_1.html", {
        "rows": rows,
        "today": dt_date.today().strftime("%Y-%m-%d"),
        "existing_dates": existing_dates,
        "page_obj": page_obj,
        "paginator": paginator,
        "per_page": per_page,
    })


@login_required
def station_table_1_view(request, date_str):
    if request.user.is_staff or request.user.is_superuser:
        return redirect("admin_table1_reports")

    d = _parse_date(date_str)
    has_night = _station_has_night(request.user)
    blocks = _terminal_blocks_for_station_date(request.user, d, force_new=False)

    blocks_ctx = []
    any_total = False

    for b in blocks:
        day_obj = StationDailyTable1.objects.filter(station_user=request.user, date=d, shift="day", block=b).first()
        night_obj = StationDailyTable1.objects.filter(station_user=request.user, date=d, shift="night", block=b).first()
        total_obj = StationDailyTable1.objects.filter(station_user=request.user, date=d, shift="total", block=b).first()

        if total_obj is not None:
            any_total = True

        blocks_ctx.append({
            "b": b,
            "day_obj": day_obj,
            "night_obj": night_obj,
            "total_obj": total_obj,
        })

    if not any_total:
        return redirect("station_table_1_edit", date_str=date_str)

    return render(request, "station_table_1_create.html", {
        "date": d,
        "blocks_ctx": blocks_ctx,
        "station_name": request.user.username,
        "mode": "view",
        "TABLE1_FIELDS": TABLE1_FIELDS,
        "is_new": False,
        "status": has_night,
    })


@login_required
def station_table_1_edit(request, date_str):
    if request.user.is_staff or request.user.is_superuser:
        return redirect("admin_table1_reports")

    has_night = _station_has_night(request.user)
    d_url = _parse_date(date_str)

    force_new = (request.GET.get("new") == "1")
    error = None

    def get_obj(shift: str, block: int):
        return StationDailyTable1.objects.filter(
            station_user=request.user, date=d_url, shift=shift, block=block
        ).first()

    blocks = _terminal_blocks_for_station_date(request.user, d_url, force_new=force_new)

    if force_new:
        blocks_ctx = [{"b": b, "day_obj": None, "night_obj": None, "total_obj": None} for b in blocks]
        is_new = True
    else:
        blocks_ctx = []
        any_total = False

        for b in blocks:
            day_obj = get_obj("day", b)
            night_obj = get_obj("night", b)
            total_obj = get_obj("total", b)

            if total_obj is not None:
                any_total = True

            blocks_ctx.append({
                "b": b,
                "day_obj": day_obj,
                "night_obj": night_obj,
                "total_obj": total_obj,
            })

        is_new = (not any_total)

    if request.method == "POST":
        posted_date_str = (request.POST.get("date") or "").strip()
        d_form = _parse_date(posted_date_str) if posted_date_str else d_url
        d_save = d_form if is_new else d_url

        if is_new and StationDailyTable1.objects.filter(
            station_user=request.user, date=d_save, shift="total"
        ).exists():
            error = f"Отчёт за {d_save.strftime('%d.%m.%Y')} уже существует. Выберите другую дату."
            return render(request, "station_table_1_create.html", {
                "date": d_save,
                "blocks_ctx": blocks_ctx,
                "station_name": request.user.username,
                "mode": "edit",
                "TABLE1_FIELDS": TABLE1_FIELDS,
                "is_new": True,
                "error": error,
                "status": has_night,
            })

        posted_blocks = set()
        for k in request.POST.keys():
            if not k.startswith("b"):
                continue
            try:
                head = k.split("__", 1)[0]
                if head.startswith("b"):
                    n = int(head[1:])
                    if n > 0:
                        posted_blocks.add(n)
            except Exception:
                continue

        blocks_to_save = sorted(posted_blocks)
        if not blocks_to_save:
            blocks_to_save = [1]

        existing_blocks = set(
            StationDailyTable1.objects
            .filter(station_user=request.user, date=d_save)
            .values_list("block", flat=True)
            .distinct()
        )
        blocks_to_delete = existing_blocks - set(blocks_to_save)
        if blocks_to_delete:
            StationDailyTable1.objects.filter(
                station_user=request.user,
                date=d_save,
                block__in=blocks_to_delete,
            ).delete()

        for b in blocks_to_save:
            day_data = {}
            night_data = {}
            total_data = {}

            term_name = (request.POST.get(f"b{b}__terminal__name") or "").strip()
            k_key = f"b{b}__common__k_podache_so_st"
            k_val = _read_int(request.POST.get(k_key))

            for key, _label in TABLE1_FIELDS:
                if key == "k_podache_so_st":
                    continue
                day_data[key] = _read_int(request.POST.get(f"b{b}__day__{key}"))
                if has_night:
                    night_data[key] = _read_int(request.POST.get(f"b{b}__night__{key}"))
                else:
                    night_data[key] = 0

            day_data["k_podache_so_st"] = k_val
            if has_night:
                night_data["k_podache_so_st"] = k_val

            day_data[TERMINAL_NAME_KEY] = term_name
            if has_night:
                night_data[TERMINAL_NAME_KEY] = term_name

            for key, _label in TABLE1_FIELDS:
                if key == "k_podache_so_st":
                    total_data[key] = k_val
                    continue
                if key == "income_daily":
                    continue
                total_data[key] = int(day_data.get(key, 0)) + (int(night_data.get(key, 0)) if has_night else 0)

            total_data[TERMINAL_NAME_KEY] = term_name

            for key, _label in TABLE1_FIELDS:
                if key in ("k_podache_so_st", "income_daily"):
                    continue
                manual_raw = (request.POST.get(f"b{b}__total__{key}") or "").strip()
                if manual_raw != "":
                    total_data[key] = _read_int(manual_raw)

            income_auto = 0
            for key, _label in TABLE1_FIELDS:
                if key in ("income_daily", "k_podache_so_st"):
                    continue
                income_auto += int(total_data.get(key, 0) or 0)

            income_manual_raw = (request.POST.get(f"b{b}__total__income_daily") or "").strip()
            total_data["income_daily"] = _read_int(income_manual_raw) if income_manual_raw != "" else income_auto

            StationDailyTable1.objects.update_or_create(
                station_user=request.user, date=d_save, shift="day", block=b,
                defaults={"data": day_data}
            )
            if has_night:
                StationDailyTable1.objects.update_or_create(
                    station_user=request.user, date=d_save, shift="night", block=b,
                    defaults={"data": night_data}
                )
            else:
                StationDailyTable1.objects.filter(
                    station_user=request.user, date=d_save, shift="night", block=b
                ).delete()

            StationDailyTable1.objects.update_or_create(
                station_user=request.user, date=d_save, shift="total", block=b,
                defaults={"data": total_data}
            )

        if request.POST.get("submit_report") == "1":
            now = timezone.now()
            StationDailyTable1.objects.filter(
                station_user=request.user,
                date=d_save,
                shift="total",
            ).update(submitted_at=now)

        return redirect("station_table_1_list")

    return render(request, "station_table_1_create.html", {
        "date": d_url,
        "blocks_ctx": blocks_ctx,
        "station_name": request.user.username,
        "mode": "edit",
        "TABLE1_FIELDS": TABLE1_FIELDS,
        "is_new": is_new,
        "error": error,
        "status": has_night,
    })

@login_required
def station_table_1_delete(request, date_str):
    if request.user.is_staff or request.user.is_superuser:
        return redirect("admin_table1_reports")

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    d = _parse_date(date_str)
    StationDailyTable1.objects.filter(station_user=request.user, date=d).delete()
    return redirect("station_table_1_list")


# =========================
# TABLE 2 DEFINITIONS
# =========================

TABLE2_ROWS = [
    (1,  "Прибыло всего:",               "П",    "r01_total", "r01_ktk"),
    (2,  "В том числе груж.всего",       "ПГ",   "r02_total", "r02_ktk"),
    (3,  "Из них под сортировку",        "ПГС",  "r03_total", "r03_ktk"),
    (4,  "Порожних",                     "ПП",   "r04_total", "r04_ktk"),
    (5,  "Поступило из ремонта",         "ИН",   "r05_total", "r05_ktk"),
    (6,  "Поступило соб.(приват)",       "ПС",   "r06_total", "r06_ktk"),
    (7,  "Поступило новых",              "Н",    "r07_total", "r07_ktk"),
    (8,  "Принято на баланс",            "ПБ",   "r08_total", "r08_ktk"),
    (9,  "Изъято из резерва",            "ПР",   "r09_total", "r09_ktk"),
    (10, "Изъято из запаса",             "ПЗ",   "r10_total", "r10_ktk"),
    (11, "Завоз автотранспортом",        "ПТ",   "r11_total", "r11_ktk"),
    (12, "Погружено-всего:",             "С",    "r12_total", "r12_ktk"),
    (13, "В том числе груженых",         "СГ",   "r13_total", "r13_ktk"),
    (14, "порожних",                     "СП",   "r14_total", "r14_ktk"),
    (15, "Поступило в ремонт",           "СН",   "r15_total", "r15_ktk"),
    (16, "Выбыло соб.(приват)",          "СС",   "r16_total", "r16_ktk"),
    (17, "Исключено",                    "ИН",   "r17_total", "r17_ktk"),
    (18, "Передано на баланс",           "СБ",   "r18_total", "r18_ktk"),
    (19, "Отставлено в резерве",         "СР",   "r19_total", "r19_ktk"),
    (20, "Отставание в запасе",          "СЗ",   "r20_total", "r20_ktk"),
    (21, "Вывоз автотранспортом",        "СТ",   "r21_total", "r21_ktk"),
    (22, "Загружено",                    "З",    R22_G_TOTAL, R22_G_KTK),
    (35, "Порожние",                     "p",    "r22p_total", "r22p_ktk"),
    (23, "Разгружено",                   "Р",    "r23_total", "r23_ktk"),
    (24, "Наличие на КП",                "В",    "r24_total", "r24_ktk"),
    (25, "В рабочем парке на пл",        "ВР",   "r25_total", "r25_ktk"),
    (26, "В том числе груженых",         "ВРГ",  "r26_total", "r26_ktk"),
    (27, "Из них под сортировку",        "ВРГС", "r27_total", "r27_ktk"),
    (28, "Готовых к отправлению",        "ВРГО", "r28_total", "r28_ktk"),
    (29, "К вывозу",                     "ВРВ",  "r29_total", "r29_ktk"),
    (30, "Порожних",                     "ВРП",  "r30_total", "r30_ktk"),
    (31, "в нерабочем парке",            "ВН",   "r31_total", "r31_ktk"),
    (32, "В том числе в резерве",        "ВНР",  "r32_total", "r32_ktk"),
    (33, "Неисправных",                  "ВНИ",  "r33_total", "r33_ktk"),
    (34, "Наличие в запасе",             "КЗ",   "r34_total", "r34_ktk"),
]

TABLE2_BOTTOM_FIELDS = {
    "income": "income_daily",

    "vygr_wag_total": "vygr_wag_total",
    "vygr_wag_ktk": "vygr_wag_ktk",
    "vygr_tonn": "vygr_tonn",
    "vygr_income": "vygr_income",

    "pogr_wag_total": "pogr_wag_total",
    "pogr_wag_ktk": "pogr_wag_ktk",
    "pogr_tonn": "pogr_tonn",
    "pogr_income": "pogr_income",

    "os_wag_total": "os_wag_total",
    "os_wag_ktk": "os_wag_ktk",
    "os_tonn": "os_tonn",
    "os_income": "os_income",

    "cargo_name": "cargo_name",
    "cargo_volume": "cargo_volume",
    "cargo_income": "cargo_income",

    # backward compatibility old keys
    "kp_fp_capacity": "kp_fp_capacity",
    "kp_fp_fact": "kp_fp_fact",
    "kp_fp_free": "kp_fp_free",
    "kp_uus_capacity": "kp_uus_capacity",
    "kp_uus_fact": "kp_uus_fact",
    "kp_uus_free": "kp_uus_free",

    "kp_ready_send": "kp_ready_send",
    "kp_ready_autocar": "kp_ready_autocar",
    "kp_ready_send_capacity": "kp_ready_send_capacity",
    "kp_ready_send_fact": "kp_ready_send_fact",
    "kp_ready_send_free": "kp_ready_send_free",
    "kp_ready_autocar_capacity": "kp_ready_autocar_capacity",
    "kp_ready_autocar_fact": "kp_ready_autocar_fact",
    "kp_ready_autocar_free": "kp_ready_autocar_free",
}


# =========================
# TABLE2 KP ROW HELPERS
# =========================

def _clean_sector_rows(rows):
    """
    rows = [{"name":..,"capacity":..,"fact":..,"free":..}, ...]
    keeps at least one row
    """
    cleaned = []

    for row in (rows or []):
        name = str((row or {}).get("name") or "").strip()
        cap = _int0((row or {}).get("capacity"))
        fact = _int0((row or {}).get("fact"))
        free = _int0((row or {}).get("free"))

        cleaned.append({
            "name": name,
            "capacity": cap,
            "fact": fact,
            "free": free,
        })

    if not cleaned:
        cleaned = [{
            "name": "",
            "capacity": 0,
            "fact": 0,
            "free": 0,
        }]

    return cleaned


def _read_sector_rows_from_post(request):
    """
    Reads:
      kp_sector_name[]
      kp_sector_capacity[]
      kp_sector_fact[]
      kp_sector_free[]
    """
    names = request.POST.getlist("kp_sector_name[]")
    caps = request.POST.getlist("kp_sector_capacity[]")
    facts = request.POST.getlist("kp_sector_fact[]")
    frees = request.POST.getlist("kp_sector_free[]")

    max_len = max(len(names), len(caps), len(facts), len(frees), 1)
    rows = []

    for i in range(max_len):
        name = (names[i] if i < len(names) else "").strip()
        cap = _read_int(caps[i] if i < len(caps) else "")
        fact = _read_int(facts[i] if i < len(facts) else "")
        free = _read_int(frees[i] if i < len(frees) else "")

        if name == "" and cap == 0 and fact == 0 and free == 0:
            continue

        rows.append({
            "name": name,
            "capacity": cap,
            "fact": fact,
            "free": free,
        })

    return _clean_sector_rows(rows)


def _sum_sector_rows(rows):
    rows = _clean_sector_rows(rows)
    return {
        "capacity": sum(_int0(r.get("capacity")) for r in rows),
        "fact": sum(_int0(r.get("fact")) for r in rows),
        "free": sum(_int0(r.get("free")) for r in rows),
    }


def _table2_sector_rows(data: dict):
    """
    New format:
      data["kp_sector_rows"] = [
        {"name":"...", "capacity":1, "fact":2, "free":3}
      ]

    Backward compatibility:
      old keys kp_fp_rows / kp_uus_rows
      or old summary keys kp_fp_capacity ... kp_uus_free
    """
    data = data or {}

    rows = data.get("kp_sector_rows")
    if isinstance(rows, list) and rows:
        return _clean_sector_rows(rows)

    converted = []

    old_fp_rows = data.get("kp_fp_rows")
    if isinstance(old_fp_rows, list):
        for row in old_fp_rows:
            converted.append({
                "name": "ФТТ",
                "capacity": _int0((row or {}).get("capacity")),
                "fact": _int0((row or {}).get("fact")),
                "free": _int0((row or {}).get("free")),
            })

    old_uus_rows = data.get("kp_uus_rows")
    if isinstance(old_uus_rows, list):
        for row in old_uus_rows:
            converted.append({
                "name": "УЛС",
                "capacity": _int0((row or {}).get("capacity")),
                "fact": _int0((row or {}).get("fact")),
                "free": _int0((row or {}).get("free")),
            })

    if converted:
        return _clean_sector_rows(converted)

    fp_has_any = any(_int0(data.get(k)) != 0 for k in ("kp_fp_capacity", "kp_fp_fact", "kp_fp_free"))
    uus_has_any = any(_int0(data.get(k)) != 0 for k in ("kp_uus_capacity", "kp_uus_fact", "kp_uus_free"))

    if fp_has_any or uus_has_any:
        fallback_rows = []
        fallback_rows.append({
            "name": "ФТТ",
            "capacity": _int0(data.get("kp_fp_capacity")),
            "fact": _int0(data.get("kp_fp_fact")),
            "free": _int0(data.get("kp_fp_free")),
        })
        fallback_rows.append({
            "name": "УЛС",
            "capacity": _int0(data.get("kp_uus_capacity")),
            "fact": _int0(data.get("kp_uus_fact")),
            "free": _int0(data.get("kp_uus_free")),
        })
        return _clean_sector_rows(fallback_rows)

    return _clean_sector_rows([])


@login_required
def station_table_2_list(request):
    if request.user.is_staff or request.user.is_superuser:
        return redirect("admin_table2_reports")

    qs = (
        StationDailyTable2.objects
        .filter(station_user=request.user)
        .order_by("-date")
    )

    page_number = request.GET.get("page", 1)
    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(page_number)

    rows = [{
        "date": obj.date,
        "year": obj.date.year,
        "submitted_at": getattr(obj, "submitted_at", None),
    } for obj in page_obj.object_list]

    existing_dates = set(qs.values_list("date", flat=True))

    return render(request, "station_table_2.html", {
        "rows": rows,
        "page_obj": page_obj,
        "paginator": paginator,
        "today": dt_date.today().strftime("%Y-%m-%d"),
        "existing_dates": existing_dates,
    })


@login_required
def station_table_2_view(request, date_str):
    if request.user.is_staff or request.user.is_superuser:
        return redirect("admin_table2_reports")

    d = _parse_date(date_str)
    obj = StationDailyTable2.objects.filter(station_user=request.user, date=d).first()
    data = (obj.data or {}) if obj else {}

    return render(request, "station_table_2_create.html", {
        "date": d,
        "obj": obj,
        "table2_data": data,
        "station_name": request.user.username,
        "rows_def": TABLE2_ROWS,
        "mode": "view",
        "bottom": TABLE2_BOTTOM_FIELDS,
        "sector_rows": _table2_sector_rows(data),
        "r22_keys": {
            "g_total": R22_G_TOTAL, "g_ktk": R22_G_KTK,
            "p_total": R22_P_TOTAL, "p_ktk": R22_P_KTK,
        }
    })


@login_required
def station_table_2_edit(request, date_str):
    if request.user.is_staff or request.user.is_superuser:
        return redirect("admin_table2_reports")

    d_url = _parse_date(date_str)
    force_new = (request.GET.get("new") == "1")

    if force_new:
        obj = None
        is_new = True
    else:
        obj = StationDailyTable2.objects.filter(station_user=request.user, date=d_url).first()
        is_new = (obj is None)

    error = None

    if request.method == "POST":
        posted_date_str = (request.POST.get("date") or "").strip()
        d_form = _parse_date(posted_date_str) if posted_date_str else d_url
        d_save = d_form if is_new else d_url

        if is_new and StationDailyTable2.objects.filter(station_user=request.user, date=d_save).exists():
            error = f"Отчёт за {d_save.strftime('%d.%m.%Y')} уже существует. Выберите другую дату."
            return render(request, "station_table_2_create.html", {
                "date": d_save,
                "obj": None,
                "table2_data": {},
                "station_name": request.user.username,
                "rows_def": TABLE2_ROWS,
                "mode": "edit",
                "bottom": TABLE2_BOTTOM_FIELDS,
                "is_new": True,
                "error": error,
                "sector_rows": [{
                    "name": "",
                    "capacity": 0,
                    "fact": 0,
                    "free": 0,
                }],
                "r22_keys": {
                    "g_total": R22_G_TOTAL, "g_ktk": R22_G_KTK,
                    "p_total": R22_P_TOTAL, "p_ktk": R22_P_KTK,
                }
            })

        data = {}

        for _n, _label, _code, k_total, k_ktk in TABLE2_ROWS:
            data[k_total] = _read_int_post(request, k_total)
            data[k_ktk] = _read_int_post(request, k_ktk)

        data[R22_P_TOTAL] = _read_int_post(request, R22_P_TOTAL)
        data[R22_P_KTK]   = _read_int_post(request, R22_P_KTK)

        data[TABLE2_BOTTOM_FIELDS["income"]] = _read_int_post(request, TABLE2_BOTTOM_FIELDS["income"])

        int_keys = [
            "vygr_wag_total", "vygr_wag_ktk", "vygr_tonn", "vygr_income",
            "pogr_wag_total", "pogr_wag_ktk", "pogr_tonn", "pogr_income",
            "os_wag_total", "os_wag_ktk", "os_tonn", "os_income",
            "cargo_volume", "cargo_income",
            "kp_ready_send", "kp_ready_autocar",
            "kp_ready_send_capacity", "kp_ready_send_fact", "kp_ready_send_free",
            "kp_ready_autocar_capacity", "kp_ready_autocar_fact", "kp_ready_autocar_free",
        ]
        for k in int_keys:
            data[TABLE2_BOTTOM_FIELDS[k]] = _read_int_post(request, TABLE2_BOTTOM_FIELDS[k])

        data[TABLE2_BOTTOM_FIELDS["cargo_name"]] = (request.POST.get(TABLE2_BOTTOM_FIELDS["cargo_name"]) or "").strip()

        # NEW dynamic sector rows
        sector_rows = _read_sector_rows_from_post(request)
        sector_sum = _sum_sector_rows(sector_rows)

        data["kp_sector_rows"] = sector_rows

        # compatibility totals
        data["kp_sector_capacity_total"] = sector_sum["capacity"]
        data["kp_sector_fact_total"] = sector_sum["fact"]
        data["kp_sector_free_total"] = sector_sum["free"]

        StationDailyTable2.objects.update_or_create(
            station_user=request.user,
            date=d_save,
            defaults={"data": data}
        )

        return redirect("station_table_2_list")

    table2_data = (obj.data or {}) if obj else {}

    return render(request, "station_table_2_create.html", {
        "date": d_url,
        "obj": obj,
        "table2_data": table2_data,
        "station_name": request.user.username,
        "rows_def": TABLE2_ROWS,
        "mode": "edit",
        "bottom": TABLE2_BOTTOM_FIELDS,
        "is_new": is_new,
        "error": error,
        "sector_rows": _table2_sector_rows(table2_data),
        "r22_keys": {
            "g_total": R22_G_TOTAL, "g_ktk": R22_G_KTK,
            "p_total": R22_P_TOTAL, "p_ktk": R22_P_KTK,
        }
    })


@login_required
def station_table_2_delete(request, date_str):
    if request.user.is_staff or request.user.is_superuser:
        return redirect("admin_table2_reports")

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    d = _parse_date(date_str)
    StationDailyTable2.objects.filter(station_user=request.user, date=d).delete()
    return redirect("station_table_2_list")


@staff_required
def promote_station(request, pk):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    station = get_object_or_404(StationProfile, id=pk)
    station.status = not station.status
    station.save()
    return redirect("admin_stations")


# =========================
# ADMIN: TABLE 1
# =========================
from django.db.models import Count, Q, Max

def admin_table1_reports(request):
    return render(request, "admin_table1_reports.html")


@staff_required
@require_GET
def admin_table1_reports_json(request):
    all_stations = _get_all_stations()
    all_station_ids = [sid for sid, _ in all_stations]
    total_count = len(all_stations)

    qs_dates = (
        StationDailyTable1.objects
        .filter(shift="total", station_user_id__in=all_station_ids)
        .values("date")
        .annotate(
            submitted_count=Count(
                "station_user_id",
                filter=Q(submitted_at__isnull=False),
                distinct=True
            ),
            last_submitted_at=Max("submitted_at"),
        )
        .order_by("-date")
    )

    per_page = _read_int(request.GET.get("per_page")) or 10
    if per_page not in (5, 10, 20, 50):
        per_page = 10

    paginator = Paginator(qs_dates, per_page)
    page_number = request.GET.get("page") or 1
    page_obj = paginator.get_page(page_number)

    items = []
    for r in page_obj.object_list:
        d = r["date"]
        submitted_count = int(r.get("submitted_count") or 0)
        last_submitted_at = r.get("last_submitted_at")

        items.append({
            "date": d.strftime("%Y-%m-%d") if d else None,
            "year": d.year if d else None,
            "month": d.month if d else None,
            "submitted_at": last_submitted_at.strftime("%d.%m.%Y %H:%M") if last_submitted_at else None,
            "is_submitted": bool(last_submitted_at),
            "submitted_count": submitted_count,
            "not_submitted_count": max(total_count - submitted_count, 0),
            "total_count": total_count,
        })

    return JsonResponse({
        "ok": True,
        "items": items,
        "today": dt_date.today().strftime("%Y-%m-%d"),
        "pagination": {
            "page": page_obj.number,
            "per_page": per_page,
            "num_pages": paginator.num_pages,
            "total_items": paginator.count,
            "has_previous": page_obj.has_previous(),
            "has_next": page_obj.has_next(),
            "previous_page_number": page_obj.previous_page_number() if page_obj.has_previous() else None,
            "next_page_number": page_obj.next_page_number() if page_obj.has_next() else None,
            "start_index": page_obj.start_index() if paginator.count else 0,
            "end_index": page_obj.end_index() if paginator.count else 0,
        }
    })


@staff_required
@require_GET
def admin_table1_status_detail(request, date_str):
    d = _parse_date(date_str)

    all_stations = _get_all_stations()
    all_station_ids = [sid for sid, _ in all_stations]

    station_name_map = {
        sp.user_id: (sp.station_name or getattr(sp.user, "username", str(sp.user)))
        for sp in StationProfile.objects.select_related("user").filter(user_id__in=all_station_ids)
    }

    sent_rows = list(
        StationDailyTable1.objects
        .filter(
            date=d,
            shift="total",
            station_user_id__in=all_station_ids,
            submitted_at__isnull=False,
        )
        .values("station_user_id")
        .annotate(last=Max("submitted_at"))
        .order_by("station_user_id")
    )

    sent_map = {x["station_user_id"]: x["last"] for x in sent_rows}

    submitted = []
    not_submitted = []

    for sid, username in all_stations:
        station_name = station_name_map.get(sid, username)
        last_dt = sent_map.get(sid)

        if last_dt:
            submitted.append({
                "name": station_name,
                "submitted_at": last_dt.strftime("%d.%m.%Y %H:%M"),
            })
        else:
            not_submitted.append({
                "name": station_name
            })

    return JsonResponse({
        "ok": True,
        "date": d.strftime("%d.%m.%Y"),
        "submitted_count": len(submitted),
        "not_submitted_count": len(not_submitted),
        "submitted": submitted,
        "not_submitted": not_submitted,
    })


@staff_required
@require_GET
def admin_table2_status_detail(request, date_str):
    d = _parse_date(date_str)

    all_stations = _get_all_stations()
    all_station_ids = [sid for sid, _ in all_stations]

    station_name_map = {
        sp.user_id: (sp.station_name or getattr(sp.user, "username", str(sp.user)))
        for sp in StationProfile.objects.select_related("user").filter(user_id__in=all_station_ids)
    }

    sent_rows = list(
        StationDailyTable2.objects
        .filter(
            date=d,
            shift="total",
            station_user_id__in=all_station_ids,
            submitted_at__isnull=False,
        )
        .values("station_user_id")
        .annotate(last=Max("submitted_at"))
        .order_by("station_user_id")
    )

    sent_map = {x["station_user_id"]: x["last"] for x in sent_rows}

    submitted = []
    not_submitted = []

    for sid, username in all_stations:
        station_name = station_name_map.get(sid, username)
        last_dt = sent_map.get(sid)

        if last_dt:
            submitted.append({
                "name": station_name,
                "submitted_at": last_dt.strftime("%d.%m.%Y %H:%M"),
            })
        else:
            not_submitted.append({
                "name": station_name
            })

    return JsonResponse({
        "ok": True,
        "date": d.strftime("%d.%m.%Y"),
        "submitted_count": len(submitted),
        "not_submitted_count": len(not_submitted),
        "submitted": submitted,
        "not_submitted": not_submitted,
    })



def _apply_itogo_rules(data: dict) -> dict:
    d = dict(data or {})

    blocks = [
        ("vygr",      "ft", "cont", "kr", "pv", "proch"),
        ("pod_vygr",  "ft", "cont", "kr", "pv", "proch"),
        ("pogr",      "ft", "cont", "kr", "pv", "proch"),
        ("pod_pogr",  "ft", "cont", "kr", "pv", "proch"),
    ]

    for prefix, k_ft, k_cont, k_kr, k_pv, k_proch in blocks:
        ft = _int0(d.get(f"{prefix}_{k_ft}"))
        kr = _int0(d.get(f"{prefix}_{k_kr}"))
        pv = _int0(d.get(f"{prefix}_{k_pv}"))
        proch = _int0(d.get(f"{prefix}_{k_proch}"))
        cont = _int0(d.get(f"{prefix}_{k_cont}"))

        d[f"{prefix}_itogo"] = ft + kr + pv + proch
        d[f"{prefix}_itogo_kon"] = cont

    return d


def _table1_part_field_name():
    try:
        field_names = {f.name for f in StationDailyTable1._meta.get_fields()}
    except Exception:
        return None

    candidates = [
        "part", "table_part", "table_no", "table_num",
        "table_index", "table_idx",
        "block", "block_no", "group", "group_no",
        "pack", "pack_no", "set_no", "form_no",
    ]
    for nm in candidates:
        if nm in field_names:
            return nm
    return None


def _sum_dicts(dicts):
    out = {}
    if not dicts:
        return out

    keys = set()
    for d in dicts:
        keys |= set((d or {}).keys())

    for k in keys:
        if k in ("k_podache_so_st", TERMINAL_NAME_KEY):
            val = ""
            for d in dicts:
                v = (d or {}).get(k, "")
                if v not in ("", None):
                    val = v
                    break
            out[k] = val
            continue

        s = 0
        for d in dicts:
            s += _int0((d or {}).get(k))
        out[k] = s

    return out


def _get_table1_shift_data_for_admin(user, d, shift: str):
    part_field = _table1_part_field_name()
    qs = StationDailyTable1.objects.filter(station_user=user, date=d, shift=shift)

    if part_field:
        objs = list(qs.order_by(part_field))
        data_list = [(o.data or {}) for o in objs]
        return _sum_dicts(data_list)

    obj = qs.first()
    return (obj.data or {}) if obj else {}


def _station_display_name(user):
    try:
        sp = StationProfile.objects.select_related("user").get(user=user)
    except StationProfile.DoesNotExist:
        return getattr(user, "username", str(user))

    candidates = [
        "station_name", "name", "title", "display_name", "short_name",
        "lc_name", "center_name", "department_name",
    ]
    for f in candidates:
        v = getattr(sp, f, None)
        if v:
            return str(v)

    for rel in ["station", "lc", "center", "department", "branch"]:
        obj = getattr(sp, rel, None)
        if obj:
            for f in ["station_name", "title", "short_name"]:
                v = getattr(obj, f, None)
                if v:
                    return str(v)
            return str(obj)

    return getattr(user, "profilestation", str(user.profilestation.station_name))


@staff_required
def admin_table1_report_view(request, date_str):
    d = _parse_date(date_str)

    User = get_user_model()
    users = (
        User.objects
        .exclude(is_staff=True)
        .exclude(is_superuser=True)
        .order_by("username")
    )

    FIELDS = [
        "podano_lc", "k_podache_so_st",
        "vygr_ft", "vygr_cont", "vygr_kr", "vygr_pv", "vygr_proch", "vygr_itogo", "vygr_itogo_kon",
        "pod_vygr_ft", "pod_vygr_cont", "pod_vygr_kr", "pod_vygr_pv", "pod_vygr_proch", "pod_vygr_itogo", "pod_vygr_itogo_kon",
        "uborka",
        "pogr_ft", "pogr_cont", "pogr_kr", "pogr_pv", "pogr_proch", "pogr_itogo_kon",
        "pod_pogr_ft", "pod_pogr_cont", "pod_pogr_kr", "pod_pogr_pv", "pod_pogr_proch", "pod_pogr_itogo_kon",
        "income_daily",
    ]

    def _to_int(v):
        if v in (None, "", "—", "-", "–"):
            return 0
        if isinstance(v, str):
            v = v.replace("\xa0", "").replace(" ", "").replace(",", "")
        try:
            return int(float(v))
        except Exception:
            return 0

    def _sum_into(dst: dict, src: dict):
        for k in FIELDS:
            dst[k] = dst.get(k, 0) + _to_int(src.get(k, 0))

    station_list = []

    for u in users:
        try:
            sp = StationProfile.objects.select_related("user").get(user=u)
        except StationProfile.DoesNotExist:
            continue

        has_night = bool(sp.status)

        sent = StationDailyTable1.objects.filter(
            station_user=u,
            date=d,
            submitted_at__isnull=False,
        ).exists()
        if not sent:
            continue

        blocks = _terminal_blocks_for_station_date(u, d, force_new=False)

        terminals = []
        for b in blocks:
            day_obj = StationDailyTable1.objects.filter(station_user=u, date=d, shift="day", block=b).first()
            night_obj = StationDailyTable1.objects.filter(station_user=u, date=d, shift="night", block=b).first()

            day_raw = (day_obj.data or {}) if day_obj else {}
            night_raw = (night_obj.data or {}) if (night_obj and has_night) else {}

            day_data = _apply_itogo_rules(day_raw)
            night_data = _apply_itogo_rules(night_raw) if has_night else {}

            total_data = {}
            if has_night:
                total_data = {k: 0 for k in FIELDS}
                _sum_into(total_data, day_data)
                _sum_into(total_data, night_data)
                total_data = _apply_itogo_rules(total_data)

            term_name = (
                (day_raw or {}).get(TERMINAL_NAME_KEY)
                or (night_raw or {}).get(TERMINAL_NAME_KEY)
                or ""
            )

            terminals.append({
                "block": b,
                "terminal_name": term_name,
                "day_data": day_data,
                "night_data": night_data,
                "total_data": total_data,
            })

        sum_total = {k: 0 for k in FIELDS}
        for t in terminals:
            _sum_into(sum_total, t["day_data"])
            if has_night:
                _sum_into(sum_total, t["night_data"])
        sum_total = _apply_itogo_rules(sum_total)

        blocks_url = ""
        if has_night:
            blocks_url = reverse(
                "admin_table1_station_blocks",
                kwargs={"date_str": d.strftime("%Y-%m-%d"), "user_id": u.id},
            )

        station_list.append({
            "name": _station_display_name(u),
            "login": getattr(u, "username", ""),
            "user_id": u.id,
            "status": has_night,
            "blocks_url": blocks_url,
            "terminals": terminals,
            "sum_total": sum_total,
        })

    station_list.sort(key=lambda x: (x["name"] or "").lower())
    grand_total = {k: 0 for k in FIELDS}

    for st in station_list:
        _sum_into(grand_total, st.get("sum_total") or {})

    grand_total = _apply_itogo_rules(grand_total)

    return render(request, "admin_table1_report_view.html", {
        "date": d,
        "stations": station_list,
        "fields": TABLE1_FIELDS,
        "grand_total": grand_total,
    })


# =========================
# ADMIN: TABLE 2
# =========================

STATION_TO_DEPT = {
    "Ташкент": 1,
    "Коканд": 2,
    "Бухара": 3,
    "Ургенч": 4,
    "Питняк": 5,
    "Нукус": 6,
}
DEPT_ORDER = [1, 2, 3, 4, 5, 6]


def _station_name(u):
    return getattr(u, "username", str(u))


from django.db.models import Max, Count, Q
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.views.decorators.http import require_GET
from django.shortcuts import render
from accounts.models import StationProfile
from .models import StationDailyTable2


@staff_required
def admin_table2_reports(request):
    return render(request, "admin_table2_reports.html")


@staff_required
@require_GET
def admin_table2_reports_json(request):
    all_stations = _get_all_stations()
    all_station_ids = [sid for sid, _ in all_stations]
    total_count = len(all_stations)

    qs_dates = (
        StationDailyTable2.objects
        .filter(station_user_id__in=all_station_ids)
        .values("date")
        .annotate(
            submitted_count=Count(
                "station_user_id",
                filter=Q(submitted_at__isnull=False),
                distinct=True,
            ),
            last_submitted_at=Max("submitted_at"),
        )
        .order_by("-date")
    )

    per_page = _read_int(request.GET.get("per_page")) or 10
    if per_page not in (5, 10, 20, 50):
        per_page = 10

    paginator = Paginator(qs_dates, per_page)
    page_number = request.GET.get("page") or 1
    page_obj = paginator.get_page(page_number)

    items = []
    for r in page_obj.object_list:
        d = r["date"]
        submitted_count = int(r.get("submitted_count") or 0)
        last_submitted_at = r.get("last_submitted_at")

        items.append({
            "date": d.strftime("%Y-%m-%d") if d else None,
            "year": d.year if d else None,
            "month": d.month if d else None,
            "submitted_at": last_submitted_at.strftime("%d.%m.%Y %H:%M") if last_submitted_at else None,
            "is_submitted": bool(last_submitted_at),
            "submitted_count": submitted_count,
            "not_submitted_count": max(total_count - submitted_count, 0),
            "total_count": total_count,
        })

    return JsonResponse({
        "ok": True,
        "items": items,
        "today": dt_date.today().strftime("%Y-%m-%d"),
        "pagination": {
            "page": page_obj.number,
            "per_page": per_page,
            "num_pages": paginator.num_pages,
            "total_items": paginator.count,
            "has_previous": page_obj.has_previous(),
            "has_next": page_obj.has_next(),
            "previous_page_number": page_obj.previous_page_number() if page_obj.has_previous() else None,
            "next_page_number": page_obj.next_page_number() if page_obj.has_next() else None,
            "start_index": page_obj.start_index() if paginator.count else 0,
            "end_index": page_obj.end_index() if paginator.count else 0,
        }
    })


@staff_required
@require_GET
def admin_table2_status_detail(request, date_str):
    d = _parse_date(date_str)

    all_stations = _get_all_stations()
    all_station_ids = [sid for sid, _ in all_stations]

    station_name_map = {
        sp.user_id: (sp.station_name or getattr(sp.user, "username", str(sp.user)))
        for sp in StationProfile.objects.select_related("user").filter(user_id__in=all_station_ids)
    }

    sent_rows = list(
        StationDailyTable2.objects
        .filter(
            date=d,
            station_user_id__in=all_station_ids,
            submitted_at__isnull=False,
        )
        .values("station_user_id")
        .annotate(last=Max("submitted_at"))
        .order_by("station_user_id")
    )

    sent_map = {x["station_user_id"]: x["last"] for x in sent_rows}

    submitted = []
    not_submitted = []

    for sid, username in all_stations:
        station_name = station_name_map.get(sid, username)
        last_dt = sent_map.get(sid)

        if last_dt:
            submitted.append({
                "name": station_name,
                "submitted_at": last_dt.strftime("%d.%m.%Y %H:%M"),
            })
        else:
            not_submitted.append({
                "name": station_name
            })

    return JsonResponse({
        "ok": True,
        "date": d.strftime("%d.%m.%Y"),
        "submitted_count": len(submitted),
        "not_submitted_count": len(not_submitted),
        "submitted": submitted,
        "not_submitted": not_submitted,
    })


@staff_required
def admin_table2_day(request, date_str):
    d = _parse_date(date_str)

    qs = StationDailyTable2.objects.filter(date=d).select_related("station_user")
    cnt = qs.count()
    last = qs.aggregate(last=Max("submitted_at"))["last"]

    return render(request, "admin_table2_day.html", {
        "date": d,
        "cnt": cnt,
        "last": last,
    })


@staff_required
def admin_table2_view(request, date_str):
    return redirect("admin_table2_station_pick", date_str=date_str)


@staff_required
def admin_table2_graph(request, date_str):
    d = _parse_date(date_str)

    objs = (
        StationDailyTable2.objects
        .filter(date=d, submitted_at__isnull=False)
        .select_related("station_user")
        .exclude(station_user__is_staff=True)
        .exclude(station_user__is_superuser=True)
        .order_by("station_user__username")
    )

    stations = [{
        "name":get_object_or_404(StationProfile, user__username= _station_name(o.station_user)).station_name,
        "data": o.data or {},
    } for o in objs]

    if not stations:
        return render(request, "admin_table2_graph.html", {
            "date": d,
            "stations": [],
            "grid": [],
        })

    all_keys = []
    for _n, _label, _code, k_total, k_ktk in TABLE2_ROWS:
        all_keys.append(k_total)
        all_keys.append(k_ktk)

    road_data = {k: 0 for k in all_keys}
    for st in stations:
        data = st["data"]
        for k in all_keys:
            road_data[k] += _dget(data, k, 0)

    stations_plus = stations + [{"name": "Дорога", "data": road_data}]

    grid = []
    for n, label, code, k_total, k_ktk in TABLE2_ROWS:
        row = {"n": n, "label": label, "code": code, "cells": []}
        for st in stations_plus:
            row["cells"].append({
                "total": _dget(st["data"], k_total, 0),
                "ktk": _dget(st["data"], k_ktk, 0),
            })
        grid.append(row)

    return render(request, "admin_table2_graph.html", {
        "date": d,
        "stations": stations_plus,
        "grid": grid,
    })


@staff_required
def admin_table2_layout(request, date_str):
    d = _parse_date(date_str)

    objs = (
        StationDailyTable2.objects
        .filter(date=d, submitted_at__isnull=False)
        .select_related("station_user")
        .exclude(station_user__is_staff=True)
        .exclude(station_user__is_superuser=True)
        .order_by("station_user__username")
    )

    def empty_bucket():
        return {
            "work_cont": 0, "work_kr": 0,
            "pogr_cont": 0, "pogr_kr": 0,
            "vygr_cont": 0, "vygr_kr": 0,
            "vygr_tuk": 0,
            "site_cont": 0, "site_kr": 0,
            "to_export_cont": 0, "to_export_kr": 0,
            "ready_cont": 0, "ready_kr": 0,
            "empty_cont": 0, "empty_kr": 0,
            "sort_cont": 0, "sort_kr": 0,
        }

    KEY = {
        "arr_total": "r01_total",
        "work_total": "r24_total", "work_ktk": "r24_ktk",
        "pogr_total": "r12_total", "pogr_ktk": "r12_ktk",
        "vygr_total": "r23_total", "vygr_ktk": "r23_ktk",
        "site_total": "r25_total", "site_ktk": "r25_ktk",
        "to_export_total": "r29_total", "to_export_ktk": "r29_ktk",
        "ready_total": "r28_total", "ready_ktk": "r28_ktk",
        "empty_total": "r30_total", "empty_ktk": "r30_ktk",
        "sort_total": "r27_total", "sort_ktk": "r27_ktk",
    }

    def add_pair(bucket, data, total_key, ktk_key, out_total, out_ktk):
        bucket[out_total] += _dget(data, total_key, 0)
        bucket[out_ktk]   += _dget(data, ktk_key, 0)

    cols = []
    buckets = {}

    for o in objs:
        u = o.station_user
        col_key = f"u{u.id}"
        col_title =get_object_or_404(StationProfile, user__username=_station_name(u)).station_name 

        if col_key not in buckets:
            cols.append({"key": col_key, "title": col_title})
            buckets[col_key] = empty_bucket()

        data = o.data or {}
        b = buckets[col_key]

        add_pair(b, data, KEY["work_total"], KEY["work_ktk"], "work_cont", "work_kr")
        add_pair(b, data, KEY["pogr_total"], KEY["pogr_ktk"], "pogr_cont", "pogr_kr")
        add_pair(b, data, KEY["vygr_total"], KEY["vygr_ktk"], "vygr_cont", "vygr_kr")
        add_pair(b, data, KEY["site_total"], KEY["site_ktk"], "site_cont", "site_kr")
        add_pair(b, data, KEY["to_export_total"], KEY["to_export_ktk"], "to_export_cont", "to_export_kr")
        add_pair(b, data, KEY["ready_total"], KEY["ready_ktk"], "ready_cont", "ready_kr")
        add_pair(b, data, KEY["empty_total"], KEY["empty_ktk"], "empty_cont", "empty_kr")
        add_pair(b, data, KEY["sort_total"], KEY["sort_ktk"], "sort_cont", "sort_kr")

    road_key = "road"
    buckets[road_key] = empty_bucket()

    for c in cols:
        b = buckets[c["key"]]
        road = buckets[road_key]
        for k in (
            "work_cont","work_kr",
            "pogr_cont","pogr_kr",
            "vygr_cont","vygr_kr",
            "site_cont","site_kr",
            "to_export_cont","to_export_kr",
            "ready_cont","ready_kr",
            "empty_cont","empty_kr",
            "sort_cont","sort_kr",
        ):
            road[k] += int(b.get(k, 0) or 0)

    for o in objs:
        data = o.data or {}
        buckets[road_key]["vygr_tuk"] += _dget(data, KEY["arr_total"], 0)

    cols.append({"key": road_key, "title": "Дорога"})

    return render(request, "admin_table2_layout.html", {
        "date": d,
        "cols": cols,
        "buckets": buckets,
    })

@login_required(login_url='login')
def admin_table2_station_pick(request, date_str):
    d = _parse_date(date_str)

    qs = (
        StationDailyTable2.objects
        .filter(date=d, submitted_at__isnull=False)
        .select_related("station_user")
        .exclude(station_user__is_staff=True)
        .exclude(station_user__is_superuser=True)
        .order_by("station_user__username")
    )

    stations = [{
        "user_id": o.station_user_id,
        "name":get_object_or_404(StationProfile, user__username=_station_name(o.station_user)).station_name ,
        "submitted_at": o.submitted_at,
    } for o in qs]

    return render(request, "admin_table2_station_pick.html", {
        "date": d,
        "stations": stations,
    })


@staff_required
def admin_table2_station_view(request, date_str, user_id: int):
    d = _parse_date(date_str)

    obj = get_object_or_404(
        StationDailyTable2.objects.select_related("station_user"),
        date=d,
        station_user_id=user_id,
        submitted_at__isnull=False,
    )
    data = obj.data or {}

    return render(request, "station_table_2_create.html", {
        "date": d,
        "obj": obj,
        "table2_data": data,
        "station_name":get_object_or_404(StationProfile, user__username=_station_name(obj.station_user)).station_name ,
        "rows_def": TABLE2_ROWS,
        "mode": "view",
        "bottom": TABLE2_BOTTOM_FIELDS,
        "sector_rows": _table2_sector_rows(data),
        "r22_keys": {
            "g_total": R22_G_TOTAL, "g_ktk": R22_G_KTK,
            "p_total": R22_P_TOTAL, "p_ktk": R22_P_KTK,
        }
    })


def _get_all_stations():
    User = get_user_model()
    qs = User.objects.exclude(is_staff=True).exclude(is_superuser=True)

    try:
        qs2 = qs.filter(station_profile__isnull=False)
        return list(qs2.values_list("id", "username").order_by("username"))
    except FieldError:
        pass

    return list(qs.values_list("id", "username").order_by("username"))


# =========================
# ADMIN: EXPORT TABLE 1 EXCEL
# =========================

@staff_required
def admin_table1_export_excel(request, date_str):
    d = _parse_date(date_str)

    User = get_user_model()
    users = (
        User.objects
        .exclude(is_staff=True)
        .exclude(is_superuser=True)
        .order_by("username")
    )

    station_list = []
    for u in users:
        try:
            sp = StationProfile.objects.get(user=u)
        except StationProfile.DoesNotExist:
            continue

        sent = StationDailyTable1.objects.filter(
            station_user=u,
            date=d,
            shift="total",
            submitted_at__isnull=False,
        ).exists()
        if not sent:
            continue

        has_night = bool(sp.status)

        day_data_raw = _get_table1_shift_data_for_admin(u, d, "day")
        night_data_raw = _get_table1_shift_data_for_admin(u, d, "night") if has_night else {}
        total_data_raw = _get_table1_shift_data_for_admin(u, d, "total")

        day_data = _apply_itogo_rules(day_data_raw or {})
        night_data = _apply_itogo_rules(night_data_raw or {})
        total_data = _apply_itogo_rules(total_data_raw or {})

        day_income = _int0(day_data.get("income_daily"))
        night_income = _int0(night_data.get("income_daily")) if has_night else 0
        total_data["income_daily"] = day_income + night_income

        station_list.append({
            "name": u.username,
            "day": day_data,
            "night": night_data,
            "total": total_data,
            "status": has_night,
        })

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.cell.cell import MergedCell

    wb = Workbook()
    ws = wb.active
    ws.title = f"Table1 {d.strftime('%d.%m.%Y')}"

    thin = Side(style="thin", color="99A3B3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    bold = Font(bold=True)
    bold_big = Font(bold=True, size=13)
    hdr_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    vtxt = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)

    def set_cell(r, c, value=None, *, font=None, fill=None, align=None, b=border):
        cell = ws.cell(row=r, column=c)
        if value is not None:
            cell.value = value
        cell.border = b
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        return cell

    def F(hex6):
        return PatternFill("solid", fgColor=("FF" + hex6.upper()))

    fill_red_col   = F("F3D6D6")
    fill_green_hdr = F("DFF4DF")
    fill_green2_hdr= F("CFEEDF")
    fill_yellow_hdr= F("F3E3B2")
    fill_blue_hdr  = F("D9F2F9")
    fill_blue2_hdr = F("C7ECF3")
    fill_gray_hdr  = F("E7EEF7")
    fill_total_row = F("FFF2CC")
    fill_title     = F("F5F7FB")

    COLS = [
        ("podano_lc", "LMga berildi"),
        ("k_podache_so_st", "St’dan berishga"),

        ("vygr_ft", "фт"),
        ("vygr_cont", "конт."),
        ("vygr_kr", "кр"),
        ("vygr_pv", "пв"),
        ("vygr_proch", "boshqa"),
        ("vygr_itogo", "jami"),
        ("vygr_itogo_kon", "jami kon"),

        ("pod_vygr_ft", "фт"),
        ("pod_vygr_cont", "конт."),
        ("pod_vygr_kr", "кр"),
        ("pod_vygr_pv", "пв"),
        ("pod_vygr_proch", "boshqa"),
        ("pod_vygr_itogo", "jami"),
        ("pod_vygr_itogo_kon", "jami kon"),

        ("uborka", "Yig‘ishtirish"),

        ("pogr_ft", "фт"),
        ("pogr_cont", "конт."),
        ("pogr_kr", "кр"),
        ("pogr_pv", "пв"),
        ("pogr_proch", "boshqa"),
        ("pogr_itogo_kon", "jami kon"),

        ("pod_pogr_ft", "фт"),
        ("pod_pogr_cont", "конт."),
        ("pod_pogr_kr", "кр"),
        ("pod_pogr_pv", "пв"),
        ("pod_pogr_proch", "boshqa"),
        ("pod_pogr_itogo_kon", "jami kon"),

        ("income_daily", "sutkalik daromad"),
    ]

    col_name = 1
    col_shift = 2
    last_col = 2 + len(COLS)

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    set_cell(
        r, 1,
        f'Оперативная информация (Таблица 1) — {d.strftime("%d.%m.%Y")}',
        font=bold_big,
        fill=fill_title,
        align=center,
    )
    ws.row_dimensions[r].height = 24

    r1 = 2
    r2 = 3
    ws.row_dimensions[r1].height = 34
    ws.row_dimensions[r2].height = 110

    ws.merge_cells(start_row=r1, start_column=col_name, end_row=r2, end_column=col_name)
    ws.merge_cells(start_row=r1, start_column=col_shift, end_row=r2, end_column=col_shift)
    set_cell(r1, col_name, "LM nomi", font=hdr_font, align=center)
    set_cell(r1, col_shift, "Smena", font=hdr_font, align=center)

    ws.merge_cells(start_row=r1, start_column=3, end_row=r2, end_column=3)
    ws.merge_cells(start_row=r1, start_column=4, end_row=r2, end_column=4)
    set_cell(r1, 3, "LMga berildi", font=hdr_font, fill=fill_red_col, align=vtxt)
    set_cell(r1, 4, "St’dan berishga", font=hdr_font, fill=fill_red_col, align=vtxt)

    def merge_group(title, c1, c2, fill):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r1, end_column=c2)
        set_cell(r1, c1, title, font=hdr_font, fill=fill, align=center)
        for cc in range(c1, c2 + 1):
            set_cell(r2, cc, font=hdr_font, fill=fill, align=vtxt)

    merge_group("Tushirish", 5, 11, fill_green_hdr)
    merge_group("Tushirishda", 12, 18, fill_green2_hdr)

    ws.merge_cells(start_row=r1, start_column=19, end_row=r2, end_column=19)
    set_cell(r1, 19, "Yig‘ishtirish", font=hdr_font, fill=fill_yellow_hdr, align=vtxt)

    merge_group("Yuklash", 20, 25, fill_blue_hdr)
    merge_group("Yuklashda", 26, 31, fill_blue2_hdr)

    ws.merge_cells(start_row=r1, start_column=32, end_row=r2, end_column=32)
    set_cell(r1, 32, "sutkalik daromad", font=hdr_font, fill=fill_gray_hdr, align=vtxt)

    for excel_col, (key, lbl) in enumerate(COLS[2:], start=5):
        if 5 <= excel_col <= 11:
            fill = fill_green_hdr
        elif 12 <= excel_col <= 18:
            fill = fill_green2_hdr
        elif 20 <= excel_col <= 25:
            fill = fill_blue_hdr
        elif 26 <= excel_col <= 31:
            fill = fill_blue2_hdr
        else:
            fill = None
        set_cell(r2, excel_col, lbl, font=hdr_font, fill=fill, align=vtxt)

    ws.freeze_panes = "C4"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    for cc in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(cc)].width = 10

    row_idx = 4

    def safe_int(v):
        try:
            return int(v or 0)
        except (TypeError, ValueError):
            return 0

    def write_shift_row(rw, shift_label, data, is_total=False):
        set_cell(
            rw, 2, shift_label,
            font=bold if is_total else None,
            align=center,
            fill=fill_total_row if is_total else None
        )

        col = 3
        for key, _lbl in COLS:
            val = (data or {}).get(key, 0)

            if key == "k_podache_so_st":
                out = val if val is not None else ""
            else:
                out = safe_int(val)

            set_cell(
                rw, col, out,
                font=bold if is_total else None,
                align=center,
                fill=fill_total_row if is_total else None
            )
            col += 1

        set_cell(rw, 1, None, align=left, fill=fill_total_row if is_total else None)

    for st in station_list:
        has_night = bool(st["status"])
        span = 3 if has_night else 2

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + span - 1, end_column=1)
        set_cell(row_idx, 1, st["name"], font=bold, align=left)

        write_shift_row(row_idx, "kun", st["day"], is_total=False)
        row_idx += 1

        if has_night:
            write_shift_row(row_idx, "tun", st["night"], is_total=False)
            row_idx += 1

        write_shift_row(row_idx, "jami", st["total"], is_total=True)
        row_idx += 1

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)

    filename = f"table1_like_site_{d.strftime('%Y-%m-%d')}.xlsx"
    resp = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@staff_required
def admin_table1_station_blocks(request, date_str, user_id: int):
    d = _parse_date(date_str)

    User = get_user_model()
    u = get_object_or_404(User, id=user_id)

    sp = get_object_or_404(StationProfile, user=u)
    has_night = bool(sp.status)

    blocks = _terminal_blocks_for_station_date(u, d, force_new=False)

    def get_obj(shift: str, block: int):
        return StationDailyTable1.objects.filter(
            station_user=u, date=d, shift=shift, block=block
        ).first()

    blocks_ctx = []
    for b in blocks:
        blocks_ctx.append({
            "b": b,
            "day_obj": get_obj("day", b),
            "night_obj": get_obj("night", b),
            "total_obj": get_obj("total", b),
        })

    day_sum = _apply_itogo_rules(_get_table1_shift_data_for_admin(u, d, "day") or {})
    night_sum = _apply_itogo_rules(_get_table1_shift_data_for_admin(u, d, "night") or {}) if has_night else {}
    total_sum = _apply_itogo_rules(_get_table1_shift_data_for_admin(u, d, "total") or {}) if has_night else {}

    if has_night:
        total_sum["income_daily"] = _int0(day_sum.get("income_daily")) + _int0(night_sum.get("income_daily"))

    return render(request, "admin_table1_station_blocks.html", {
        "date": d,
        "station_name": u.username,
        "user_id": u.id,
        "fields": TABLE1_FIELDS,
        "blocks_ctx": blocks_ctx,
        "sum_day": day_sum,
        "sum_night": night_sum,
        "sum_total": total_sum,
        "status": has_night,
    })




import json
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from django.contrib.auth.decorators import login_required
from django.db.models import Exists, OuterRef
from django.templatetags.static import static

from .models import Notification, NotificationRead



def _safe_user_name(user):
    full_name = f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip()
    if full_name:
        return full_name
    username = getattr(user, "username", "") or ""
    if username:
        return username
    return f"User {user.id}"


def _safe_avatar_url(notification):
    try:
        if notification.avatar and hasattr(notification.avatar, "url"):
            return notification.avatar.url
    except Exception:
        pass

    try:
        creator = notification.created_by
        if creator:
            profile = getattr(creator, "profile", None)
            if profile and getattr(profile, "photo", None) and hasattr(profile.photo, "url"):
                return profile.photo.url
    except Exception:
        pass

    return static("images/admin-bot.png")


@require_GET
@login_required
def notifications_latest(request):

    """
    User/admin uchun oxirgi aktiv habarnoma.
    Admin uchun:
      - xabar
      - kimlar o'qiganini live ko'rsatish
    User uchun:
      - xabar
      - unread holat
    """
    latest_qs = Notification.objects.filter(is_active=True)


    read_subq = NotificationRead.objects.filter(
        user=request.user,
        notification=OuterRef("pk")
    )

    notif = latest_qs.annotate(
        user_has_read=Exists(read_subq)
    ).order_by("-created_at").select_related("created_by").first()

    if not notif:
        return JsonResponse({"ok": True, "notification": None, "unread": False})

    unread = not bool(getattr(notif, "user_has_read", False))

    read_events = []
    if request.user.is_staff or request.user.is_superuser:
        latest_reads = (
            NotificationRead.objects
            .filter(notification=notif)
            .select_related("user")
            .order_by("-read_at")[:20]
        )

        read_events = [
            {
                "user_id": r.user_id,
                "user_name": _safe_user_name(r.user),
                "read_at": timezone.localtime(r.read_at).strftime("%d.%m.%Y %H:%M:%S"),
            }
            for r in latest_reads
        ]

    return JsonResponse({
        "ok": True,
        "notification": {
            "id": notif.id,
            "message": notif.message,
            "created_at": timezone.localtime(notif.created_at).strftime("%d.%m.%Y %H:%M:%S"),
            "created_by_name": _safe_user_name(notif.created_by) if notif.created_by else "Admin",
            "avatar_url": _safe_avatar_url(notif),
        },
        "unread": unread,
        "read_events": read_events,
    })


@require_POST
@login_required
def notifications_ack(request):
    try:
        body = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        body = {}

    notif_id = body.get("id")
    if not notif_id:
        return JsonResponse({"ok": False, "detail": "notification_id_required"}, status=400)

    notif = Notification.objects.filter(id=notif_id, is_active=True).first()
    if not notif:
        return JsonResponse({"ok": False, "detail": "notification_not_found"}, status=404)

    obj, created = NotificationRead.objects.get_or_create(
        user=request.user,
        notification=notif
    )

    return JsonResponse({
        "ok": True,
        "marked": True,
        "created": created,
        "id": notif.id,
        "user_name": _safe_user_name(request.user),
        "read_at": timezone.localtime(obj.read_at).strftime("%d.%m.%Y %H:%M:%S"),
    })


@require_POST
@login_required
def notifications_send(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({"ok": False, "detail": "forbidden"}, status=403)

    try:
        body = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        body = {}

    msg = (body.get("message") or "").strip()
    if not msg:
        return JsonResponse({"ok": False, "detail": "empty_message"}, status=400)

    notif = Notification.objects.create(
        message=msg,
        created_by=request.user,
        is_active=True,
    )

    return JsonResponse({
        "ok": True,

        "notification": {
            "id": notif.id,
            "message": notif.message,
            "created_at": timezone.localtime(notif.created_at).strftime("%d.%m.%Y %H:%M:%S"),
            "created_by_name": _safe_user_name(request.user),
            "avatar_url": _safe_avatar_url(notif),
        }
    })


